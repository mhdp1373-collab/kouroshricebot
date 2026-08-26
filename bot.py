"""
ربات پیام‌رسان بله (Bale) مدیریت مستندات بارنامه - پی‌بار
نسخه ۴ — با منوی Reply ثابت پایین صفحه

توجه‌های مهم برای اجرا روی بله:
──────────────────────────────
1. توکن ربات را از @Bot_Father داخل خودِ پیام‌رسان بله بگیرید (نه تلگرام)
   و در متغیر محیطی BOT_TOKEN قرار دهید.
2. کتابخانه‌ی python-telegram-bot به‌صورت پیش‌فرض به سرور تلگرام وصل می‌شود؛
   در تابع main() آدرس پایه به سرور بله (tapi.bale.ai) تغییر داده شده است
   — این روش دقیقاً همان چیزی است که خودِ تیم بله در نمونه‌کدهای رسمی‌شان
   (bale-bot-samples) برای استفاده از python-telegram-bot پیشنهاد داده‌اند.
3. API بله زیرمجموعه‌ای از API تلگرام و «تا حد زیادی» سازگار با آن است، اما
   ۱۰۰٪ یکسان نیست. مواردی که ممکن است نیاز به تست/تنظیم داشته باشند:
   - برخی حالت‌های parse_mode (Markdown/HTML) ممکن است در بله رفتار
     کمی متفاوتی داشته باشند.
   - لینک‌های ویژه‌ی تلگرام مثل tg://user?id=... در بله کار نمی‌کنند؛
     به همین دلیل در این نسخه به‌جای لینک کلیک‌پذیر، آیدی راننده به‌صورت
     متن قابل‌کپی (`کد`) نمایش داده می‌شود.
   - محدودیت حجم فایل/عکس و نرخ ارسال پیام ممکن است با تلگرام فرق داشته باشد.
4. پیشنهاد می‌شود قبل از استفاده‌ی نهایی، کل سناریوها (بارگزاری، تایید،
   عدم تایید، کسری بار، پاسخ راننده) را یک‌بار کامل روی بله تست کنید.
"""

import os
import io
import re
import json
import uuid
import hashlib
import hmac
import base64
import secrets
import time
import asyncio
import logging
from types import SimpleNamespace
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
from telegram import (
    Update, InlineKeyboardButton, InlineKeyboardMarkup,
    ReplyKeyboardMarkup, ReplyKeyboardRemove, KeyboardButton
)
from telegram.ext import (
    Application, CommandHandler, MessageHandler, CallbackQueryHandler,
    ConversationHandler, filters, ContextTypes, TypeHandler, ApplicationHandlerStop
)
from dotenv import load_dotenv
from fastapi import FastAPI, HTTPException, Header, Query, Body, File, UploadFile, Form
from fastapi.responses import HTMLResponse, StreamingResponse
import uvicorn
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

load_dotenv()

# ─── زمان به وقت تهران ───
# سرور (Railway) معمولاً روی UTC اجراست؛ همه‌ی تاریخ/ساعت‌های ذخیره‌شده در دیتابیس و
# نمایش داده‌شده در ربات/داشبورد باید به وقت تهران باشند، نه وقت سرور.
TEHRAN_TZ = ZoneInfo("Asia/Tehran")

def now_tehran() -> datetime:
    """زمان فعلی به وقت تهران (بدون tzinfo، تا با رشته‌های ذخیره‌شده در دیتابیس سازگار بماند)"""
    return datetime.now(TEHRAN_TZ).replace(tzinfo=None)

def now_str(with_seconds: bool = False) -> str:
    fmt = "%Y-%m-%d %H:%M:%S" if with_seconds else "%Y-%m-%d %H:%M"
    return now_tehran().strftime(fmt)

def _gregorian_to_jalali(gy: int, gm: int, gd: int):
    """تبدیل تاریخ میلادی به شمسی (الگوریتم استاندارد، بدون نیاز به کتابخانه‌ی خارجی)"""
    g_days_in_month = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    gy2 = gy + 1 if gm > 2 else gy
    days = (
        355666 + (365 * gy) + ((gy2 + 3) // 4) - ((gy2 + 99) // 100) + ((gy2 + 399) // 400)
        + gd + sum(g_days_in_month[:gm - 1])
    )
    jy = -1595 + (33 * (days // 12053))
    days %= 12053
    jy += 4 * (days // 1461)
    days %= 1461
    if days > 365:
        jy += (days - 1) // 365
        days = (days - 1) % 365
    if days < 186:
        jm = 1 + days // 31
        jd = 1 + (days % 31)
    else:
        jm = 7 + (days - 186) // 30
        jd = 1 + ((days - 186) % 30)
    return jy, jm, jd

def to_jalali(date_str: str) -> str:
    """رشته‌ی تاریخ ذخیره‌شده (میلادی، مثلاً «2026-08-25 14:30») را به نمایش شمسی تبدیل می‌کند.
    برای نمایش تاریخ/ساعت به راننده و ادمین داخل خودِ ربات استفاده می‌شود."""
    if not date_str or date_str == "-":
        return date_str or "-"
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})(?:[ T](\d{2}):(\d{2}))?", str(date_str))
    if not m:
        return date_str
    gy, gm, gd, hh, mi = m.groups()
    jy, jm, jd = _gregorian_to_jalali(int(gy), int(gm), int(gd))
    out = f"{jy}/{jm:02d}/{jd:02d}"
    if hh is not None:
        out += f" {hh}:{mi}"
    return out

# ─── مسیر ذخیره‌سازی دائمی ───
# روی Railway (و بیشتر PaaSها)، فضای پیش‌فرض کانتینر ephemeral است: با هر دیپلوی/ری‌استارت
# پاک می‌شود. برای جلوگیری از پاک شدن database.json و bot.log، یک Volume دائمی بساز و
# مسیر Mount آن را در متغیر محیطی DATA_DIR بگذار (مثلاً DATA_DIR=/data).
# اگر DATA_DIR تنظیم نشده باشد (مثلاً حین تست روی سیستم خودت)، فایل‌ها کنار خود کد
# ذخیره می‌شوند — دقیقاً مثل رفتار قبلی.
DATA_DIR = os.getenv("DATA_DIR", "").strip()
if DATA_DIR:
    os.makedirs(DATA_DIR, exist_ok=True)

def _data_path(filename: str) -> str:
    return os.path.join(DATA_DIR, filename) if DATA_DIR else filename

logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO,
    handlers=[
        logging.FileHandler(_data_path("bot.log"), encoding="utf-8"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

if DATA_DIR:
    logger.info(f"💾 مسیر ذخیره‌سازی دائمی (DATA_DIR) فعال است: {DATA_DIR}")
else:
    logger.warning(
        "⚠️ متغیر DATA_DIR تنظیم نشده — دیتابیس کنار خود کد ذخیره می‌شود و با هر دیپلوی "
        "روی Railway پاک خواهد شد. یک Volume دائمی بساز و DATA_DIR را به مسیر Mount آن تنظیم کن."
    )

# ─── محافظ ضد ارسال تکراری ───
# اگر پلتفرم بله (یا مکانیزم polling) یک آپدیت را بیش از یک‌بار تحویل بدهد،
# این محافظ با شناسه‌ی یکتای هر آپدیت (update_id) جلوی پردازش دوباره‌اش را می‌گیرد.
_seen_update_ids: set = set()
_seen_update_ids_order: list = []
_MAX_SEEN_UPDATE_IDS = 5000

async def dedupe_update_guard(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.update_id
    if uid in _seen_update_ids:
        logger.warning(f"⚠️ آپدیت تکراری نادیده گرفته شد — update_id={uid} (این آپدیت قبلاً پردازش شده بود)")
        raise ApplicationHandlerStop
    _seen_update_ids.add(uid)
    _seen_update_ids_order.append(uid)
    if len(_seen_update_ids_order) > _MAX_SEEN_UPDATE_IDS:
        oldest = _seen_update_ids_order.pop(0)
        _seen_update_ids.discard(oldest)

# ─── تنظیمات ───
# توکن را از @Bot_Father داخل پیام‌رسان بله بگیرید
BOT_TOKEN = os.getenv("BOT_TOKEN", "YOUR_BALE_BOT_TOKEN_HERE")
ADMIN_IDS = list(map(int, os.getenv("ADMIN_IDS", "").split(","))) if os.getenv("ADMIN_IDS") else []
STORAGE_CHANNEL_ID = os.getenv("STORAGE_CHANNEL_ID", "")
ADMIN_USERNAME = os.getenv("ADMIN_USERNAME", "@koroshrice")
DB_FILE = _data_path("database.json")

# ─── تنظیمات داشبورد وب ادمین ───
# DASHBOARD_TOKEN: دسترسی کامل (مشاهده + تایید/رد/کسری/بارگزاری/پیام)
# DASHBOARD_VIEWER_TOKEN: فقط مشاهده (اختیاری — اگر خالی باشد، نقش «مشاهده‌گر» غیرفعال است)
DASHBOARD_TOKEN = os.getenv("DASHBOARD_TOKEN", "")
DASHBOARD_VIEWER_TOKEN = os.getenv("DASHBOARD_VIEWER_TOKEN", "")
PORT = int(os.getenv("PORT", "8080"))
BOT_INSTANCE = None  # بعد از ساخته‌شدن Application مقداردهی می‌شود

# آدرس پایه‌ی API بله (به‌جای api.telegram.org)
BALE_API_BASE_URL = "https://tapi.bale.ai/"
BALE_API_FILE_URL = "https://tapi.bale.ai/file/"

# ─── مراحل مکالمه ───
WAIT_BARNAME, WAIT_DOCS = range(2)

# ─── نگاشت نام مستندات (برای سازگاری با بارنامه‌های قدیمی که با فلوی مرحله‌به‌مرحله ثبت شده‌اند) ───
DOC_NAMES = {
    "bill": "📄 اصل بارنامه",
    "origin": "🔵 حواله بار مبدأ",
    "dest": "🔴 رسید بار مقصد",
    "account": "💳 شماره حساب/شبا",
}

def doc_label(key: str, doc_info: dict = None) -> str:
    """برچسب نمایشی یک مستند. مدارک جدید برچسب خودشان را دارند (مثلاً «مدرک شماره ۲»)،
    مدارک قدیمی از نگاشت بالا خوانده می‌شوند."""
    if doc_info and doc_info.get("label"):
        return doc_info["label"]
    return DOC_NAMES.get(key, key)

# ─── متن یک‌جای راهنمای ارسال مستندات (فلوی ساده‌شده — بدون مرحله‌به‌مرحله) ───
UPLOAD_REQUIREMENTS_TEXT = (
    "📋 لطفاً موارد زیر را برای این بارنامه ارسال کنید؛ پس از ارسال همه‌ی مدارک، "
    "دکمه‌ی «✅ تایید نهایی» را جهت ارسال به شرکت بزنید:\n\n"
    "📄 اصل بارنامه\n"
    "🔵 حواله خروج مبدأ (محل بارگیری)\n"
    "🔴 رسید تخلیه مقصد\n"
    "💳 شماره حساب بانک تجارت (جهت پرداخت سریع‌تر) یا شماره شبا، به نام راننده یا شرکت باربری\n"
    "   _(اگر شماره حساب فرد دیگری را می‌فرستید، رضایت‌نامه به همراه شماره ملی صاحب حساب را هم ارسال کنید)_"
)

# ─────────── پایگاه داده ───────────

def load_db() -> dict:
    if os.path.exists(DB_FILE):
        with open(DB_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def save_db(db: dict):
    with open(DB_FILE, "w", encoding="utf-8") as f:
        json.dump(db, f, ensure_ascii=False, indent=2)

def get_barname_data(barname: str) -> dict:
    db = load_db()
    return db.get(barname, {
        "barname": barname,
        "created_at": "",
        "driver_id": 0,
        "driver_name": "",
        "documents": {},
        "message_ids": {}
    })

def save_barname_data(barname: str, data: dict):
    db = load_db()
    db[barname] = data
    save_db(db)

def add_barname_log(data: dict, event: str, actor: str = "", detail: str = ""):
    """افزودن یک رویداد به تاریخچه‌ی بارنامه (برای نمایش در داشبورد).
    این تابع فقط دیکشنری data را در حافظه تغییر می‌دهد؛ خودِ save_barname_data
    باید بلافاصله بعدش صدا زده شود تا رویداد روی دیسک ذخیره شود."""
    if "log" not in data or not isinstance(data.get("log"), list):
        data["log"] = []
    data["log"].append({
        "time": now_str(with_seconds=True),
        "event": event,
        "actor": actor,
        "detail": detail,
    })

def make_review_id() -> str:
    """شناسه یکتا برای هر درخواست بررسی (هر بار ارسال نهایی، شناسه جدید می‌گیرد)"""
    return uuid.uuid4().hex[:10]

_PERSIAN_ARABIC_DIGITS = {
    "۰": "0", "۱": "1", "۲": "2", "۳": "3", "۴": "4", "۵": "5", "۶": "6", "۷": "7", "۸": "8", "۹": "9",
    "٠": "0", "١": "1", "٢": "2", "٣": "3", "٤": "4", "٥": "5", "٦": "6", "٧": "7", "٨": "8", "٩": "9",
}

def normalize_barname_digits(text: str) -> str:
    """شماره بارنامه را فقط به رقم‌های انگلیسی تبدیل می‌کند — ارقام فارسی/عربی تبدیل و هر
    کاراکتر غیرعددی (فاصله، خط تیره، حروف و ...) حذف می‌شود."""
    converted = "".join(_PERSIAN_ARABIC_DIGITS.get(ch, ch) for ch in (text or ""))
    return "".join(ch for ch in converted if ch.isdigit())

def find_barname_by_review_id(rid: str):
    """پیدا کردن بارنامه بر اساس شناسه بررسی فعال آن"""
    db = load_db()
    for bn, data in db.items():
        if data.get("review", {}).get("id") == rid:
            return bn, data
    return None, None

def make_barname_token(barname: str) -> str:
    """شناسه کوتاه و پایدار برای هر شماره بارنامه (برای استفاده در callback_data)"""
    return hashlib.md5(barname.encode("utf-8")).hexdigest()[:12]

def find_barname_by_token(token: str):
    db = load_db()
    for bn in db.keys():
        if make_barname_token(bn) == token:
            return bn
    return None

def review_status_label(db_data: dict) -> str:
    """برچسب وضعیت آخرین بررسی یک بارنامه"""
    review = db_data.get("review", {})
    status = review.get("status")
    if status == "approved":
        return "✅ تایید شده"
    if status == "rejected":
        return "❌ تایید نشده"
    if status == "partial":
        resp = review.get("driver_response")
        if resp == "accepted":
            return "⚠️ کسری بار (راننده تایید کرد)"
        if resp == "rejected":
            return "⚠️ کسری بار (راننده تایید نکرد)"
        return "⚠️ کسری بار (در انتظار پاسخ راننده)"
    if status == "pending":
        return "🕐 در انتظار بررسی"
    return "⏳ ثبت نشده / ناقص"

# رویدادهایی که واقعاً «تصمیم/واکنش ادمین به درخواست» محسوب می‌شوند (نه اقدامات جانبی مثل افزودن عکس یا تغییر نوع محصول)
ADMIN_DECISION_EVENTS = {
    "تایید مستندات", "عدم تایید مستندات", "تایید با کسری بار",
    "تایید نهایی بارنامه (پس از کسری بار)",
}

def admin_response_times(data: dict):
    """اولین و آخرین زمان واکنش (تصمیم) ادمین به یک بارنامه را برمی‌گرداند — (اولین, آخرین) یا (None, None)"""
    times = sorted(
        e.get("time", "") for e in data.get("log", [])
        if e.get("event") in ADMIN_DECISION_EVENTS and "ادمین" in (e.get("actor") or "")
    )
    if not times:
        return None, None
    return times[0], times[-1]

def get_driver_barnames(driver_id: int):
    """همه‌ی بارنامه‌های ثبت‌شده توسط یک راننده، از دیتابیس (نه حافظه‌ی موقت)"""
    db = load_db()
    rows = []
    for barname, data in db.items():
        if data.get("driver_id") == driver_id:
            rows.append((barname, data))
    rows.sort(key=lambda x: x[1].get("created_at", ""), reverse=True)
    return rows

# ─────────── کیبوردها ───────────

def main_menu_keyboard() -> InlineKeyboardMarkup:
    """منوی اصلی راننده"""
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("🚀 شروع بارگزاری مدارک", callback_data="start_upload")],
        [InlineKeyboardButton("📋 راهنما", callback_data="show_help")],
    ])

def admin_keyboard() -> InlineKeyboardMarkup:
    """منوی اصلی ادمین"""
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("🔍 دریافت مستندات بارنامه", callback_data="admin_get")],
        [InlineKeyboardButton("📊 لیست بارنامه‌ها", callback_data="admin_list")],
        [InlineKeyboardButton("✅ بارنامه‌های تایید شده (۵ روز اخیر)", callback_data="admin_approved_list")],
        [InlineKeyboardButton("🕐 نیازمند بررسی / تایید نشده", callback_data="admin_pending_list")],
        [InlineKeyboardButton("📎 افزودن عکس/فایل به بارنامه", callback_data="admin_attach")],
        [InlineKeyboardButton("🗑 حذف بارنامه", callback_data="admin_delete")],
        [InlineKeyboardButton("📈 آمار کلی", callback_data="admin_stats")],
    ])

def driver_reply_keyboard() -> ReplyKeyboardMarkup:
    """منوی ثابت پایین صفحه برای راننده"""
    return ReplyKeyboardMarkup(
        [
            [KeyboardButton("🚀 بارگزاری مدارک"), KeyboardButton("📋 راهنما")],
            [KeyboardButton("📦 وضعیت بارنامه"), KeyboardButton("❌ لغو عملیات")],
        ],
        resize_keyboard=True,
        input_field_placeholder="از منوی پایین انتخاب کنید..."
    )

def admin_reply_keyboard() -> ReplyKeyboardMarkup:
    """منوی ثابت پایین صفحه برای ادمین"""
    return ReplyKeyboardMarkup(
        [
            [KeyboardButton("🔍 دریافت مستندات"), KeyboardButton("📊 لیست بارنامه‌ها")],
            [KeyboardButton("✅ بارنامه‌های تایید شده"), KeyboardButton("🕐 نیازمند بررسی")],
            [KeyboardButton("📎 افزودن عکس/فایل"), KeyboardButton("🗑 حذف بارنامه")],
            [KeyboardButton("📈 آمار کلی"), KeyboardButton("🏠 منوی اصلی ادمین")],
        ],
        resize_keyboard=True,
        input_field_placeholder="پنل مدیریت پی‌بار"
    )


def upload_docs_keyboard() -> ReplyKeyboardMarkup:
    """کیبورد ثابت پایین صفحه در حین ارسال مدارک — بدون دکمه شیشه‌ای زیر پیام‌ها"""
    return ReplyKeyboardMarkup(
        [
            [KeyboardButton("✅ تایید نهایی تمامی مستندات و ارسال به شرکت")],
            [KeyboardButton("❌ لغو عملیات")],
        ],
        resize_keyboard=True,
        input_field_placeholder="مدارک را ارسال کنید یا تایید نهایی بزنید"
    )

def review_keyboard(rid: str, barname: str = "") -> InlineKeyboardMarkup:
    """کیبورد بررسی مستندات برای ادمین"""
    rows = [
        [InlineKeyboardButton("✅ تأیید", callback_data=f"radm_appr_{rid}")],
        [InlineKeyboardButton("❌ عدم تأیید", callback_data=f"radm_rej_{rid}")],
        [InlineKeyboardButton("⚠️ تأیید با کسری بار", callback_data=f"radm_part_{rid}")],
    ]
    if barname:
        token = make_barname_token(barname)
        rows.append([InlineKeyboardButton(f"📦 نمایش مجدد مستندات {barname}", callback_data=f"admin_open_{token}")])
    return InlineKeyboardMarkup(rows)

def driver_partial_response_keyboard(rid: str) -> InlineKeyboardMarkup:
    """کیبورد پاسخ راننده به تایید با کسری بار"""
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("✅ مورد تایید است", callback_data=f"drv_ok_{rid}")],
        [InlineKeyboardButton("❌ مورد تایید نیست", callback_data=f"drv_no_{rid}")],
    ])

def barname_entry_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("🏠 بازگشت به منوی اصلی", callback_data="back_to_main")],
    ])

def admin_back_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("🔙 بازگشت به پنل ادمین", callback_data="admin_back")],
    ])

# ─────────── کمکی ───────────

async def go_to_main_menu(query, context, is_admin=False):
    """بازگشت به منوی اصلی"""
    context.user_data.clear()
    user_name = query.from_user.first_name

    if is_admin:
        await query.edit_message_text(
            f"🔐 *پنل مدیریت پی‌بار*\n\nسلام {user_name} عزیز! از منوی زیر انتخاب کنید:",
            parse_mode="Markdown",
            reply_markup=admin_keyboard()
        )
    else:
        await query.edit_message_text(
            f"🏠 *منوی اصلی پی‌بار*\n\nسلام {user_name} عزیز!\nبرای شروع بارگزاری مدارک دکمه زیر را بزنید:",
            parse_mode="Markdown",
            reply_markup=main_menu_keyboard()
        )

# ─────────── یادآوری تایید نهایی (اگر راننده مستندات را فرستاد ولی یادش رفت تایید نهایی بزند) ───────────

UPLOAD_REMINDER_DELAY_SECONDS = 2 * 60  # ۲ دقیقه

def _upload_reminder_job_name(chat_id: int) -> str:
    return f"upload_reminder_{chat_id}"

def _schedule_upload_reminder(context, chat_id: int, barname: str):
    """هر بار مدرک جدیدی می‌رسد، این تایمر ریست می‌شود — یعنی یادآوری دقیقاً ۲ دقیقه
    بعد از آخرین فعالیت راننده ارسال می‌شود، نه ۲ دقیقه بعد از شروع."""
    if not context.job_queue:
        return
    job_name = _upload_reminder_job_name(chat_id)
    for job in context.job_queue.get_jobs_by_name(job_name):
        job.schedule_removal()
    context.job_queue.run_once(
        _upload_reminder_callback,
        when=UPLOAD_REMINDER_DELAY_SECONDS,
        chat_id=chat_id,
        user_id=chat_id,
        data={"barname": barname},
        name=job_name,
    )

def _cancel_upload_reminder(context, chat_id: int):
    if not context.job_queue:
        return
    for job in context.job_queue.get_jobs_by_name(_upload_reminder_job_name(chat_id)):
        job.schedule_removal()

async def _upload_reminder_callback(context: ContextTypes.DEFAULT_TYPE):
    """اگر راننده هنوز داخل همون بارنامه است و تایید نهایی نزده، یادآوری می‌فرستد"""
    job = context.job
    barname = job.data.get("barname")
    # اگر کاربر تایید نهایی زده، لغو کرده، یا رفته سراغ بارنامه‌ی دیگری، user_data دیگر این بارنامه را نشان نمی‌دهد
    if context.user_data.get("barname") != barname:
        return
    session_docs = context.user_data.get("session_documents", {})
    count = len(session_docs)
    try:
        await context.bot.send_message(
            chat_id=job.chat_id,
            text=(
                f"⏰ راننده گرامی، مستندات بارنامه شماره {barname} را هنوز تایید نهایی نکرده‌اید "
                f"({count} مدرک ارسال شده).\n\n"
                "اگر مدارک شما کامل است، دکمه‌ی «✅ تایید نهایی...» پایین صفحه را بزنید تا برای شرکت ارسال شود."
            ),
            reply_markup=upload_docs_keyboard()
        )
    except Exception as e:
        logger.error(f"خطا در ارسال یادآوری بارگزاری به {job.chat_id}: {e}")

# ─────────── هندلرهای اصلی ───────────

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user = update.effective_user
    context.user_data.clear()

    if user.id in ADMIN_IDS:
        await update.message.reply_text(
            f"👋 سلام {user.first_name} عزیز!\n\n"
            "🔐 *پنل مدیریت پی‌بار*\n"
            "از منوی پایین صفحه انتخاب کنید:",
            parse_mode="Markdown",
            reply_markup=admin_reply_keyboard()
        )
        await update.message.reply_text(
            "یا از دکمه‌های زیر استفاده کنید:",
            reply_markup=admin_keyboard()
        )
        return ConversationHandler.END

    await update.message.reply_text(
        f"👋 سلام {user.first_name} عزیز!\n\n"
        "🚛 به ربات مستندات *پی‌بار* خوش آمدید.\n\n"
        "━━━━━━━━━━━━━━━━━━━━\n"
        "📋 *نحوه کار:*\n"
        "۱. دکمه شروع را بزنید\n"
        "۲. شماره بارنامه را وارد کنید\n"
        "۳. ربات یک‌به‌یک هر مستند را درخواست می‌کند\n"
        "۴. عکس بگیرید و ارسال کنید\n"
        "━━━━━━━━━━━━━━━━━━━━",
        parse_mode="Markdown",
        reply_markup=driver_reply_keyboard()
    )
    await update.message.reply_text(
        "یا از دکمه‌های زیر شروع کنید:",
        reply_markup=main_menu_keyboard()
    )
    return ConversationHandler.END


async def handle_start_upload_button(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """دکمه‌ی «🚀 بارگزاری مدارک» از منوی پایین صفحه — باید entry_point مکالمه باشد تا وضعیت درست ثبت شود"""
    await update.message.reply_text(
        "📝 لطفاً *شماره بارنامه* را وارد کنید:",
        parse_mode="Markdown",
        reply_markup=barname_entry_keyboard()
    )
    return WAIT_BARNAME


async def handle_start_upload(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """شروع فرایند بارگزاری از منوی اصلی"""
    query = update.callback_query
    await query.answer()

    await query.edit_message_text(
        "📝 لطفاً *شماره بارنامه* را وارد کنید:\n\n"
        "_(حداقل ۳ کاراکتر)_",
        parse_mode="Markdown",
        reply_markup=barname_entry_keyboard()
    )
    return WAIT_BARNAME


async def resubmit_barname(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """بارگذاری مجدد مستندات یک بارنامه رد شده، از طریق دکمه پیام عدم تایید"""
    query = update.callback_query
    await query.answer()

    token = query.data.replace("resubmit_", "")
    barname = find_barname_by_token(token)
    if not barname:
        await query.message.reply_text("⚠️ بارنامه یافت نشد.")
        return ConversationHandler.END

    existing = get_barname_data(barname)
    existing_review = existing.get("review", {})
    existing_status = existing_review.get("status")

    if existing_status == "approved":
        await query.message.reply_text(
            f"⚠️ بارنامه شماره {barname} قبلاً ارسال شده و مورد تایید قرار گرفته است، "
            f"اگر تا کنون واریز به حساب شما انجام نشده است به آیدی ادمین {ADMIN_USERNAME} پیگیری فرمایید.",
            reply_markup=driver_reply_keyboard()
        )
        return ConversationHandler.END

    if existing_status == "partial":
        driver_response = existing_review.get("driver_response")
        rid = existing_review.get("id", "")
        deduction_note = existing_review.get("deduction_note", "-")

        if driver_response == "accepted":
            await query.message.reply_text(
                f"⚠️ بارنامه شماره {barname} دارای کسری تایید شده‌ی شما می‌باشد، "
                "لطفاً تا زمان بررسی و تایید ادمین صبور باشید.",
                reply_markup=driver_reply_keyboard()
            )
        elif driver_response == "rejected":
            await query.message.reply_text(
                f"⚠️ بارنامه شماره {barname} دارای کسری می‌باشد که قبلاً توسط شما تایید نشده است، "
                f"لطفاً تا زمان بررسی و تایید ادمین صبور باشید و یا با آیدی ادمین {ADMIN_USERNAME} پیگیری نمایید.",
                reply_markup=driver_reply_keyboard()
            )
        else:
            await query.message.reply_text(
                f"⚠️ بارنامه شماره {barname} را قبلاً ارسال کرده‌اید و دارای کسری «{deduction_note}» می‌باشد.\n\n"
                "در صورت تایید، از کرایه‌ی بارنامه مبلغ آن کسر و مابقی به حساب شما واریز می‌گردد.",
                reply_markup=driver_partial_response_keyboard(rid)
            )
        return ConversationHandler.END

    # مدارک قبلی حفظ می‌شوند و با مدارک جدید ادغام می‌شوند — پاک نمی‌شوند
    context.user_data.clear()
    context.user_data["barname"] = barname
    prev_docs = dict(existing.get("documents", {}))
    context.user_data["session_documents"] = prev_docs
    existing_nums = [
        int(k.split("_", 1)[1]) for k in prev_docs
        if k.startswith("doc_") and k.split("_", 1)[1].isdigit()
    ]
    context.user_data["doc_counter"] = max(existing_nums) if existing_nums else 0

    prev_count = len(prev_docs)
    await query.message.reply_text(
        f"🔄 بارگذاری مجدد مدارک بارنامه *{barname}*\n\n"
        f"✅ {prev_count} مدرکی که قبلاً فرستاده بودید همچنان نزد ما محفوظ است و پاک نشده.\n"
        "فقط کافیست مدرکِ ناقص یا اصلاح‌شده را دوباره بفرستید — همراه با مدارک قبلی برای ادمین ارسال می‌شود.\n\n"
        f"{UPLOAD_REQUIREMENTS_TEXT}",
        parse_mode="Markdown",
        reply_markup=upload_docs_keyboard()
    )
    _schedule_upload_reminder(context, query.message.chat_id, barname)
    return WAIT_DOCS


async def handle_show_help(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """نمایش راهنما"""
    query = update.callback_query
    await query.answer()

    await query.edit_message_text(
        "📋 *راهنمای استفاده از ربات پی‌بار*\n\n"
        "━━━━━━━━━━━━━━━━━━━━\n"
        "🔹 *مستندات مورد نیاز:*\n"
        "۱. 📄 اصل بارنامه\n"
        "۲. 🔵 حواله بار مبدأ\n"
        "۳. 🔴 رسید بار مقصد\n"
        "۴. 💳 شماره حساب بانک تجارت (جهت پرداخت سریع‌تر) یا شماره شبا (به نام راننده اول یا دوم بارنامه)\n\n"
        "━━━━━━━━━━━━━━━━━━━━\n"
        "💡 *نکات مهم:*\n"
        "• عکس‌ها باید واضح و خوانا باشند\n"
        "• در هر مرحله می‌توانید برگردید\n"
        "• در صورت خطا دکمه بازگشت را بزنید\n"
        "• /cancel برای لغو کامل عملیات",
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("🚀 شروع بارگزاری", callback_data="start_upload")],
            [InlineKeyboardButton("🏠 منوی اصلی", callback_data="back_to_main")],
        ])
    )


async def back_to_main(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """بازگشت به منوی اصلی — برای راننده و ادمین"""
    query = update.callback_query
    await query.answer()
    user = query.from_user
    is_admin = user.id in ADMIN_IDS
    await go_to_main_menu(query, context, is_admin=is_admin)
    return ConversationHandler.END


async def receive_barname(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    barname = normalize_barname_digits(update.message.text)

    if not barname or len(barname) < 3:
        await update.message.reply_text(
            "⚠️ شماره بارنامه باید فقط عدد باشد (حداقل ۳ رقم).\nلطفاً دوباره وارد کنید:",
            reply_markup=barname_entry_keyboard()
        )
        return WAIT_BARNAME

    # اگر این بارنامه قبلاً تایید نهایی شده، اجازه‌ی ارسال مجدد داده نمی‌شود
    existing = get_barname_data(barname)
    existing_review = existing.get("review", {})
    existing_status = existing_review.get("status")

    if existing_status == "approved":
        await update.message.reply_text(
            f"⚠️ بارنامه شماره {barname} قبلاً ارسال شده و مورد تایید قرار گرفته است، "
            f"اگر تا کنون واریز به حساب شما انجام نشده است به آیدی ادمین {ADMIN_USERNAME} پیگیری فرمایید.",
            reply_markup=driver_reply_keyboard()
        )
        context.user_data.clear()
        return ConversationHandler.END

    if existing_status == "partial":
        driver_response = existing_review.get("driver_response")
        rid = existing_review.get("id", "")
        deduction_note = existing_review.get("deduction_note", "-")

        if driver_response == "accepted":
            await update.message.reply_text(
                f"⚠️ بارنامه شماره {barname} دارای کسری تایید شده‌ی شما می‌باشد، "
                "لطفاً تا زمان بررسی و تایید ادمین صبور باشید.",
                reply_markup=driver_reply_keyboard()
            )
        elif driver_response == "rejected":
            await update.message.reply_text(
                f"⚠️ بارنامه شماره {barname} دارای کسری می‌باشد که قبلاً توسط شما تایید نشده است، "
                f"لطفاً تا زمان بررسی و تایید ادمین صبور باشید و یا با آیدی ادمین {ADMIN_USERNAME} پیگیری نمایید.",
                reply_markup=driver_reply_keyboard()
            )
        else:
            await update.message.reply_text(
                f"⚠️ بارنامه شماره {barname} را قبلاً ارسال کرده‌اید و دارای کسری «{deduction_note}» می‌باشد.\n\n"
                "در صورت تایید، از کرایه‌ی بارنامه مبلغ آن کسر و مابقی به حساب شما واریز می‌گردد.",
                reply_markup=driver_partial_response_keyboard(rid)
            )
        context.user_data.clear()
        return ConversationHandler.END

    context.user_data["barname"] = barname
    context.user_data.pop("last_ack_message_id", None)
    # اگر این بارنامه قبلاً مدارکی داشته (رد شده یا حتی هنوز در انتظار بررسی)، آن‌ها حفظ و با مدارک جدید ادغام می‌شوند
    prev_docs = dict(existing.get("documents", {})) if existing_status in ("rejected", "pending") else {}
    context.user_data["session_documents"] = prev_docs
    existing_nums = [
        int(k.split("_", 1)[1]) for k in prev_docs
        if k.startswith("doc_") and k.split("_", 1)[1].isdigit()
    ]
    context.user_data["doc_counter"] = max(existing_nums) if existing_nums else 0

    if prev_docs:
        await update.message.reply_text(
            f"🔄 بارنامه *{barname}* شناسایی شد.\n\n"
            f"✅ {len(prev_docs)} مدرکی که قبلاً برای این بارنامه فرستاده بودید همچنان نزد ما محفوظ است.\n"
            "فقط کافیست مدرکِ ناقص یا اصلاح‌شده را دوباره بفرستید.\n\n"
            f"{UPLOAD_REQUIREMENTS_TEXT}",
            parse_mode="Markdown",
            reply_markup=upload_docs_keyboard()
        )
    else:
        await update.message.reply_text(
            f"✅ بارنامه *{barname}* ثبت شد.\n\n{UPLOAD_REQUIREMENTS_TEXT}",
            parse_mode="Markdown",
            reply_markup=upload_docs_keyboard()
        )
    _schedule_upload_reminder(context, update.effective_chat.id, barname)
    return WAIT_DOCS


def upload_ack_inline_keyboard() -> InlineKeyboardMarkup:
    """دکمه‌های زیر پیام «دریافت شد» — تا راننده همیشه یک اقدام واضح جلوی چشمش داشته باشد"""
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("✅ تایید نهایی و ارسال به شرکت", callback_data="upload_finalize_inline")],
        [InlineKeyboardButton("❌ لغو ارسال", callback_data="upload_cancel_inline")],
    ])


async def _strip_previous_ack_buttons(context, chat_id: int):
    """دکمه‌های شیشه‌ای پیام «دریافت شد» قبلی را حذف می‌کند — فقط آخرین پیام باید دکمه داشته باشد"""
    prev_msg_id = context.user_data.get("last_ack_message_id")
    if prev_msg_id:
        try:
            await context.bot.edit_message_reply_markup(chat_id=chat_id, message_id=prev_msg_id, reply_markup=None)
        except Exception:
            pass  # پیام ممکن است قبلاً بدون دکمه شده باشد یا پاک شده باشد — مشکلی نیست


async def receive_upload_item(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """دریافت هر عکس/فایل/متنی که راننده در حین ارسال مدارک بفرستد — بدون مرحله‌بندی و بدون دکمه شیشه‌ای.
    عکس/فایل به‌صورت یک مدرک شماره‌دار جدید ذخیره می‌شود؛ متن به‌عنوان شماره حساب/شبا در نظر گرفته می‌شود."""
    barname = context.user_data.get("barname")
    if not barname:
        await update.message.reply_text(
            "⚠️ لطفاً ابتدا از منوی پایین صفحه «🚀 بارگزاری مدارک» را بزنید.",
            reply_markup=driver_reply_keyboard()
        )
        return ConversationHandler.END

    session_docs = context.user_data.setdefault("session_documents", {})
    photo = update.message.photo[-1] if update.message.photo else None
    doc = update.message.document
    text = update.message.text
    chat_id = update.effective_chat.id

    if doc and not photo:
        # فقط PDF به‌عنوان «فایل» قابل قبول است — بقیه‌ی فرمت‌ها رد می‌شوند تا داشبورد بتواند نوعشان را درست تشخیص بدهد
        mime = (doc.mime_type or "").lower()
        fname = (doc.file_name or "").lower()
        is_pdf = (mime == "application/pdf") or fname.endswith(".pdf")
        if not is_pdf:
            await update.message.reply_text(
                "⚠️ فقط *عکس* یا فایل *PDF* قابل قبول است.\nلطفاً مدرک را به‌صورت عکس یا PDF ارسال کنید.",
                parse_mode="Markdown"
            )
            return WAIT_DOCS

    if photo or doc:
        context.user_data["doc_counter"] = context.user_data.get("doc_counter", 0) + 1
        n = context.user_data["doc_counter"]
        if photo:
            file_id, file_type = photo.file_id, "photo"
        else:
            file_id, file_type = doc.file_id, "pdf"

        session_docs[f"doc_{n}"] = {
            "file_id": file_id,
            "file_type": file_type,
            "text": "",
            "label": f"📎 مدرک شماره {n}",
            "uploaded_at": now_str(),
        }
        count = sum(1 for k in session_docs if k.startswith("doc_"))
        await _strip_previous_ack_buttons(context, chat_id)
        sent = await update.message.reply_text(
            f"📥 مدرک شماره {n} دریافت شد. (مجموع مدارک ارسالی: {count})\n\n"
            "⚠️ *هنوز چیزی برای شرکت ارسال نشده است.*\n"
            "اگر مدرک دیگری دارید همین الان بفرستید. وقتی همه‌ی مدارک کامل شد، حتماً "
            "«✅ تایید نهایی» را بزنید — تا آن را نزنید، مدارک شما اصلاً ارسال نمی‌شود.",
            parse_mode="Markdown",
            reply_markup=upload_ack_inline_keyboard()
        )
        context.user_data["last_ack_message_id"] = sent.message_id
    elif text and text.strip():
        session_docs["account"] = {
            "file_id": None,
            "file_type": "text",
            "text": text.strip(),
            "label": "💳 شماره حساب/شبا اعلامی",
            "uploaded_at": now_str(),
        }
        await _strip_previous_ack_buttons(context, chat_id)
        sent = await update.message.reply_text(
            f"📥 شماره حساب/شبا دریافت شد:\n`{text.strip()}`\n\n"
            "⚠️ *هنوز چیزی برای شرکت ارسال نشده است.*\n"
            "اگر مدرک دیگری دارید همین الان بفرستید. وقتی همه‌ی مدارک کامل شد، حتماً "
            "«✅ تایید نهایی» را بزنید — تا آن را نزنید، مدارک شما اصلاً ارسال نمی‌شود.",
            parse_mode="Markdown",
            reply_markup=upload_ack_inline_keyboard()
        )
        context.user_data["last_ack_message_id"] = sent.message_id
    else:
        await update.message.reply_text("⚠️ لطفاً فقط عکس، فایل PDF یا شماره حساب/شبا ارسال کنید.")
        return WAIT_DOCS

    _schedule_upload_reminder(context, chat_id, barname)
    return WAIT_DOCS


async def _do_final_confirm_upload(context, chat_id: int, user, barname: str, session_docs: dict):
    """منطق مشترک ثبت نهایی — چه از دکمه‌ی پایین صفحه صدا زده شود، چه از دکمه‌ی شیشه‌ای زیر پیام"""
    _cancel_upload_reminder(context, chat_id)

    db_data = get_barname_data(barname)
    if not db_data.get("created_at"):
        db_data["created_at"] = now_str()
    db_data["driver_id"] = user.id
    db_data["driver_name"] = user.full_name
    db_data["documents"] = session_docs
    db_data["completed_at"] = now_str()
    db_data["status"] = "completed"
    add_barname_log(
        db_data, "ارسال نهایی مستندات",
        actor=f"{user.full_name} (راننده)",
        detail=f"{len(session_docs)} مدرک ارسال شد"
    )
    save_barname_data(barname, db_data)

    if STORAGE_CHANNEL_ID and str(STORAGE_CHANNEL_ID) not in [str(a) for a in ADMIN_IDS]:
        for doc_key, doc_info in session_docs.items():
            try:
                await send_single_doc(
                    context.bot, STORAGE_CHANNEL_ID, doc_key, doc_info, barname,
                    extra_caption=f"👤 راننده: {user.full_name}"
                )
            except Exception as e:
                logger.error(f"خطا در فوروارد به کانال آرشیو: {e}")

    await dispatch_review_package(context, barname, db_data)


async def final_confirm_upload(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """دکمه‌ی «✅ تایید نهایی تمامی مستندات و ارسال به شرکت» از پنل پایین صفحه"""
    barname = context.user_data.get("barname")
    session_docs = context.user_data.get("session_documents", {})

    if not barname:
        await update.message.reply_text(
            "⚠️ لطفاً ابتدا از منوی پایین صفحه «🚀 بارگزاری مدارک» را بزنید.",
            reply_markup=driver_reply_keyboard()
        )
        return ConversationHandler.END

    if not session_docs:
        await update.message.reply_text(
            "⚠️ هنوز هیچ مدرکی ارسال نکرده‌اید.\n\nلطفاً ابتدا مدارک را بفرستید، بعد دکمه‌ی تایید نهایی را بزنید."
        )
        return WAIT_DOCS

    await _do_final_confirm_upload(context, update.effective_chat.id, update.effective_user, barname, session_docs)

    await update.message.reply_text(
        f"✅ *مدارک بارنامه {barname} با موفقیت برای شرکت ارسال شد.*\n\n"
        "نتیجه‌ی بررسی به‌زودی از طریق همین ربات به شما اطلاع داده خواهد شد.",
        parse_mode="Markdown",
        reply_markup=driver_reply_keyboard()
    )
    context.user_data.clear()
    return ConversationHandler.END


async def final_confirm_upload_inline(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """همان تایید نهایی، از دکمه‌ی شیشه‌ای زیر پیام «دریافت شد»"""
    query = update.callback_query
    barname = context.user_data.get("barname")
    session_docs = context.user_data.get("session_documents", {})

    if not barname or not session_docs:
        await query.answer("⚠️ هنوز مدرکی ثبت نشده است.", show_alert=True)
        return WAIT_DOCS

    await query.answer("در حال ارسال...")
    try:
        await query.edit_message_reply_markup(reply_markup=None)
    except Exception:
        pass

    await _do_final_confirm_upload(context, query.message.chat_id, query.from_user, barname, session_docs)

    await query.message.reply_text(
        f"✅ *مدارک بارنامه {barname} با موفقیت برای شرکت ارسال شد.*\n\n"
        "نتیجه‌ی بررسی به‌زودی از طریق همین ربات به شما اطلاع داده خواهد شد.",
        parse_mode="Markdown",
        reply_markup=driver_reply_keyboard()
    )
    context.user_data.clear()
    return ConversationHandler.END


async def cancel_upload_inline(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """دکمه‌ی «❌ لغو ارسال» زیر پیام «دریافت شد»"""
    query = update.callback_query
    await query.answer()
    try:
        await query.edit_message_reply_markup(reply_markup=None)
    except Exception:
        pass
    _cancel_upload_reminder(context, query.message.chat_id)
    context.user_data.clear()
    await query.message.reply_text(
        "❌ عملیات لغو شد.\n\nبرای شروع مجدد از منوی پایین صفحه استفاده کنید:",
        reply_markup=driver_reply_keyboard()
    )
    return ConversationHandler.END


async def send_single_doc(bot, chat_id, doc_key: str, doc_info: dict, barname: str, extra_caption: str = "", parse_mode: str = None):
    """ارسال یک مستند (عکس/فایل/متن) به یک چت مشخص"""
    label = doc_label(doc_key, doc_info)
    caption = f"{label}\n📦 بارنامه: {barname}"
    if extra_caption:
        caption += f"\n{extra_caption}"

    ftype = doc_info.get("file_type")
    if ftype == "photo":
        await bot.send_photo(chat_id=chat_id, photo=doc_info["file_id"], caption=caption, parse_mode=parse_mode)
    elif ftype in ("document", "pdf"):
        await bot.send_document(chat_id=chat_id, document=doc_info["file_id"], caption=caption, parse_mode=parse_mode)
    else:
        await bot.send_message(chat_id=chat_id, text=f"{caption}\n\n📝 {doc_info.get('text', '-')}", parse_mode=parse_mode)


async def dispatch_review_package(context, barname: str, db_data: dict) -> str:
    """ارسال پکیج مستندات و دکمه‌های بررسی برای همه‌ی ادمین‌ها؛ یک شناسه بررسی جدید تولید می‌کند"""
    rid = make_review_id()
    db_data["review"] = {
        "id": rid,
        "status": "pending",
        "reason": "",
        "deduction_note": "",
        "reviewed_by": None,
        "reviewed_at": None,
        "admin_messages": {},
    }
    add_barname_log(
        db_data, "ارسال/نمایش مجدد مستندات برای بررسی",
        detail=f"برای {len(ADMIN_IDS)} ادمین ارسال شد"
    )
    save_barname_data(barname, db_data)

    docs = db_data.get("documents", {})
    driver_id = db_data.get("driver_id")
    driver_name = db_data.get("driver_name", "-")
    driver_link = f"{driver_name} (آیدی: `{driver_id}`)" if driver_id else driver_name

    if ADMIN_IDS:
        for admin_id in ADMIN_IDS:
            try:
                # ارسال مستندات به صورت پکیج برای ادمین
                for doc_key, doc_info in docs.items():
                    await send_single_doc(
                        context.bot, admin_id, doc_key, doc_info, barname,
                        extra_caption=f"👤 فرستنده: {driver_link}",
                        parse_mode="Markdown"
                    )

                package_text = (
                    f"📦 مستندات بارنامه شماره `{barname}` توسط آیدی {driver_link} ارسال شد\n\n"
                    f"📄 تعداد مستندات: {len(docs)}\n"
                    f"🕐 زمان: {now_str()}\n\n"
                    "برای مشاهده‌ی مجدد مستندات این بارنامه، دکمه‌ی «📦 نمایش مجدد مستندات» زیر همین پیام را بزنید.\n\n"
                    "لطفاً بررسی و اقدام کنید:"
                )
                sent = await context.bot.send_message(
                    chat_id=admin_id,
                    text=package_text,
                    parse_mode="Markdown",
                    reply_markup=review_keyboard(rid, barname)
                )
                db_data = get_barname_data(barname)
                db_data["review"]["admin_messages"][str(admin_id)] = sent.message_id
                save_barname_data(barname, db_data)
            except Exception as e:
                logger.error(f"خطا در ارسال پکیج به ادمین {admin_id}: {e}")

    return rid


# ─────────── بررسی مستندات توسط ادمین ───────────

async def _finalize_review(context, barname: str, db_data: dict, status_label: str):
    """حذف دکمه‌های بررسی از پیام تمام ادمین‌ها و اطلاع‌رسانی تصمیم نهایی به آن‌ها"""
    admin_messages = db_data.get("review", {}).get("admin_messages", {})
    for admin_id_str, msg_id in admin_messages.items():
        try:
            await context.bot.edit_message_reply_markup(
                chat_id=int(admin_id_str),
                message_id=msg_id,
                reply_markup=None
            )
        except Exception as e:
            logger.error(f"خطا در حذف دکمه‌های ادمین {admin_id_str}: {e}")
        try:
            await context.bot.send_message(
                chat_id=int(admin_id_str),
                text=f"📦 بارنامه {barname}: {status_label}"
            )
        except Exception as e:
            logger.error(f"خطا در اطلاع‌رسانی به ادمین {admin_id_str}: {e}")


async def _send_partial_finalize_prompt(context, barname: str, db_data: dict, rid: str, text: str):
    """برای بارنامه‌ای که «تایید با کسری بار» شده، به همه‌ی ادمین‌ها یک پیام با دکمه‌ی
    «✅ تایید نهایی بارنامه» می‌فرستد. این دکمه در هر لحظه‌ای — چه راننده کسری را قبول کند،
    چه نکند، چه اصلاً هنوز پاسخی نداده باشد — قابل استفاده است؛ تصمیم نهایی همیشه با ادمین است.
    شناسه‌ی پیام‌های ارسالی ذخیره می‌شود تا بعد از تایید نهایی، دکمه از همه‌شان حذف شود."""
    finalize_kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("✅ تایید نهایی بارنامه", callback_data=f"radm_fin_{rid}")]
    ])
    finalize_messages = dict(db_data.get("review", {}).get("finalize_messages", {}))
    for admin_id in ADMIN_IDS:
        try:
            sent = await context.bot.send_message(chat_id=admin_id, text=text, reply_markup=finalize_kb)
            finalize_messages[str(admin_id)] = sent.message_id
        except Exception as e:
            logger.error(f"خطا در ارسال دکمه تایید نهایی به ادمین {admin_id}: {e}")

    db_data = get_barname_data(barname)  # تازه‌سازی قبل از ذخیره تا رویدادهای اخیر پاک نشوند
    db_data.setdefault("review", {})["finalize_messages"] = finalize_messages
    save_barname_data(barname, db_data)


async def review_approve(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """تایید مستندات توسط ادمین"""
    query = update.callback_query
    user = query.from_user
    if user.id not in ADMIN_IDS:
        logger.warning(f"⛔️ تلاش دسترسی غیرمجاز به دکمه ادمین توسط آیدی {user.id} (لیست ادمین‌های مجاز: {ADMIN_IDS})")
        await query.answer("⛔️ دسترسی ندارید.", show_alert=True)
        return

    rid = query.data.replace("radm_appr_", "")
    barname, db_data = find_barname_by_review_id(rid)
    if not barname or db_data.get("review", {}).get("status") != "pending":
        logger.warning(f"⚠️ درخواست بررسی نامعتبر — rid={rid} یافت‌نشد یا قبلاً بررسی شده (دیتابیس ممکن است ری‌ست شده باشد)")
        await query.answer("⚠️ این درخواست دیگر معتبر نیست یا قبلاً بررسی شده.", show_alert=True)
        return

    await query.answer("✅ تأیید شد")

    db_data["review"]["status"] = "approved"
    db_data["review"]["reviewed_by"] = user.id
    db_data["review"]["reviewed_at"] = now_str()
    add_barname_log(db_data, "تایید مستندات", actor=f"{user.first_name} (ادمین، از طریق ربات)")
    save_barname_data(barname, db_data)

    driver_id = db_data.get("driver_id")
    if driver_id:
        try:
            await context.bot.send_message(
                chat_id=driver_id,
                text=(
                    f"✅ راننده محترم مستندات ارسالی بارنامه شماره {barname} شما تایید شد "
                    "و ظرف ۲ روز کاری هزینه بارنامه به شماره حساب اعلامی شما واریز خواهد شد."
                ),
                reply_markup=driver_reply_keyboard()
            )
        except Exception as e:
            logger.error(f"خطا در اطلاع‌رسانی به راننده: {e}")

    await _finalize_review(context, barname, db_data, f"✅ تأیید شد توسط {user.first_name}")


async def review_finalize(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """تایید نهایی بارنامه‌ای که «تایید با کسری بار» شده — چه راننده کسری را پذیرفته باشد،
    چه نپذیرفته باشد، چه هنوز پاسخی نداده باشد؛ تصمیم نهایی همیشه با ادمین است."""
    query = update.callback_query
    user = query.from_user
    if user.id not in ADMIN_IDS:
        logger.warning(f"⛔️ تلاش دسترسی غیرمجاز به دکمه ادمین توسط آیدی {user.id} (لیست ادمین‌های مجاز: {ADMIN_IDS})")
        await query.answer("⛔️ دسترسی ندارید.", show_alert=True)
        return

    rid = query.data.replace("radm_fin_", "")
    barname, db_data = find_barname_by_review_id(rid)
    if not barname or db_data.get("review", {}).get("status") != "partial":
        await query.answer("⚠️ این درخواست دیگر معتبر نیست یا قبلاً نهایی شده.", show_alert=True)
        return

    await query.answer("✅ تایید نهایی شد")

    resp = db_data["review"].get("driver_response")
    resp_label = "پذیرفته بود" if resp == "accepted" else ("نپذیرفته بود" if resp == "rejected" else "هنوز پاسخی نداده بود")

    db_data["review"]["status"] = "approved"
    db_data["review"]["reviewed_by"] = user.id
    db_data["review"]["reviewed_at"] = now_str()
    add_barname_log(
        db_data, "تایید نهایی بارنامه (پس از کسری بار)",
        actor=f"{user.first_name} (ادمین، از طریق ربات)",
        detail=f"راننده کسری بار را {resp_label}"
    )
    save_barname_data(barname, db_data)

    driver_id = db_data.get("driver_id")
    if driver_id:
        try:
            await context.bot.send_message(
                chat_id=driver_id,
                text=(
                    f"✅ راننده محترم بارنامه شماره {barname} شما به‌صورت نهایی تایید شد "
                    "و ظرف ۲ روز کاری هزینه بارنامه (با احتساب کسری بار توافق‌شده) "
                    "به شماره حساب اعلامی شما واریز خواهد شد."
                ),
                reply_markup=driver_reply_keyboard()
            )
        except Exception as e:
            logger.error(f"خطا در اطلاع‌رسانی به راننده: {e}")

    # حذف دکمه‌ی «تایید نهایی» از پیام تمام ادمین‌ها
    finalize_messages = db_data.get("review", {}).get("finalize_messages", {})
    for admin_id_str, msg_id in finalize_messages.items():
        try:
            await context.bot.edit_message_reply_markup(
                chat_id=int(admin_id_str), message_id=msg_id, reply_markup=None
            )
        except Exception as e:
            logger.error(f"خطا در حذف دکمه تایید نهایی ادمین {admin_id_str}: {e}")

    await _finalize_review(context, barname, db_data, f"✅ تایید نهایی شد توسط {user.first_name} (پس از کسری بار)")


async def review_reject_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """شروع فرایند ثبت علت عدم تایید"""
    query = update.callback_query
    user = query.from_user
    if user.id not in ADMIN_IDS:
        logger.warning(f"⛔️ تلاش دسترسی غیرمجاز به دکمه ادمین توسط آیدی {user.id} (لیست ادمین‌های مجاز: {ADMIN_IDS})")
        await query.answer("⛔️ دسترسی ندارید.", show_alert=True)
        return

    rid = query.data.replace("radm_rej_", "")
    barname, db_data = find_barname_by_review_id(rid)
    if not barname or db_data.get("review", {}).get("status") != "pending":
        logger.warning(f"⚠️ درخواست بررسی نامعتبر — rid={rid} یافت‌نشد یا قبلاً بررسی شده (دیتابیس ممکن است ری‌ست شده باشد)")
        await query.answer("⚠️ این درخواست دیگر معتبر نیست یا قبلاً بررسی شده.", show_alert=True)
        return

    await query.answer()
    context.user_data["review_pending"] = {"action": "reject", "rid": rid, "barname": barname}

    await context.bot.send_message(
        chat_id=user.id,
        text=f"❌ لطفاً *علت عدم تأیید* مستندات بارنامه {barname} را بنویسید:",
        parse_mode="Markdown"
    )


async def review_partial_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """شروع فرایند ثبت توضیح تایید با کسری بار"""
    query = update.callback_query
    user = query.from_user
    if user.id not in ADMIN_IDS:
        logger.warning(f"⛔️ تلاش دسترسی غیرمجاز به دکمه ادمین توسط آیدی {user.id} (لیست ادمین‌های مجاز: {ADMIN_IDS})")
        await query.answer("⛔️ دسترسی ندارید.", show_alert=True)
        return

    rid = query.data.replace("radm_part_", "")
    barname, db_data = find_barname_by_review_id(rid)
    if not barname or db_data.get("review", {}).get("status") != "pending":
        logger.warning(f"⚠️ درخواست بررسی نامعتبر — rid={rid} یافت‌نشد یا قبلاً بررسی شده (دیتابیس ممکن است ری‌ست شده باشد)")
        await query.answer("⚠️ این درخواست دیگر معتبر نیست یا قبلاً بررسی شده.", show_alert=True)
        return

    await query.answer()
    context.user_data["review_pending"] = {"action": "partial", "rid": rid, "barname": barname}

    await context.bot.send_message(
        chat_id=user.id,
        text=(
            f"⚠️ لطفاً *مقدار کسری بار* مربوط به بارنامه {barname} را بنویسید:\n\n"
            "_(مثال: ۲ تن، یا ۳ کیسه)_"
        ),
        parse_mode="Markdown"
    )


# ─────────── پاسخ راننده به تایید با کسری بار ───────────

async def driver_partial_accept(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """راننده کسری بار را قبول می‌کند — وضعیت هنوز «تایید شد» نمی‌شود؛
    تصمیم نهایی با دکمه‌ی «تایید نهایی» در اختیار ادمین می‌ماند."""
    query = update.callback_query
    rid = query.data.replace("drv_ok_", "")
    barname, db_data = find_barname_by_review_id(rid)

    if not barname or db_data.get("review", {}).get("status") != "partial" or db_data["review"].get("driver_response"):
        await query.answer("⚠️ این درخواست دیگر معتبر نیست یا قبلاً پاسخ داده شده.", show_alert=True)
        return

    await query.answer("✅ ثبت شد")
    await query.edit_message_reply_markup(reply_markup=None)

    db_data["review"]["driver_response"] = "accepted"
    add_barname_log(db_data, "پذیرش کسری بار توسط راننده", actor=f"{query.from_user.full_name} (راننده)")
    save_barname_data(barname, db_data)

    await query.message.reply_text(
        "✅ پاسخ شما ثبت شد. پس از تایید نهایی ادمین، هزینه بارنامه واریز خواهد شد.",
        reply_markup=driver_reply_keyboard()
    )

    await _send_partial_finalize_prompt(
        context, barname, db_data, rid,
        text=(
            f"✅ راننده کسری بار بارنامه شماره {barname} را پذیرفت.\n\n"
            "برای نهایی‌شدن و اطلاع‌رسانی واریز به راننده، دکمه‌ی زیر را بزنید 👇"
        )
    )


async def driver_partial_reject_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """راننده کسری بار را قبول ندارد — شروع دریافت علت"""
    query = update.callback_query
    rid = query.data.replace("drv_no_", "")
    barname, db_data = find_barname_by_review_id(rid)

    if not barname or db_data.get("review", {}).get("status") != "partial" or db_data["review"].get("driver_response"):
        await query.answer("⚠️ این درخواست دیگر معتبر نیست یا قبلاً پاسخ داده شده.", show_alert=True)
        return

    await query.answer()
    await query.edit_message_reply_markup(reply_markup=None)

    context.user_data["driver_dispute_pending"] = {"rid": rid, "barname": barname}

    await query.message.reply_text("❌ لطفاً *علت عدم تایید* کسری بار را بنویسید:", parse_mode="Markdown")


# ─────────── پنل ادمین ───────────

async def admin_back(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """بازگشت به پنل ادمین"""
    query = update.callback_query
    await query.answer()
    context.user_data.pop("admin_action", None)

    try:
        await query.edit_message_text(
            "🔐 *پنل مدیریت پی‌بار*\n\nاز منوی زیر انتخاب کنید:",
            parse_mode="Markdown",
            reply_markup=admin_keyboard()
        )
    except Exception as e:
        # اگه پیام قابل ویرایش نبود (مثلاً پیام قدیمی/حاوی عکس بود)، یک پیام جدید می‌فرستیم
        logger.warning(f"⚠️ ویرایش پیام در admin_back ناموفق بود ({e})؛ پیام جدید ارسال می‌شود.")
        await context.bot.send_message(
            chat_id=update.effective_chat.id,
            text="🔐 *پنل مدیریت پی‌بار*\n\nاز منوی زیر انتخاب کنید:",
            parse_mode="Markdown",
            reply_markup=admin_keyboard()
        )


async def admin_get_docs(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    context.user_data["admin_action"] = "get_docs"

    await query.edit_message_text(
        "🔍 *دریافت مستندات*\n\nشماره بارنامه مورد نظر را وارد کنید:",
        parse_mode="Markdown",
        reply_markup=admin_back_keyboard()
    )


async def admin_delete_barname(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    context.user_data["admin_action"] = "delete_barname"

    await query.edit_message_text(
        "🗑 *حذف بارنامه*\n\nشماره بارنامه‌ای که می‌خواهید حذف کنید را وارد کنید:",
        parse_mode="Markdown",
        reply_markup=admin_back_keyboard()
    )


async def admin_attach_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    context.user_data["admin_action"] = "attach_barname"
    context.user_data.pop("admin_attach_barname", None)

    await query.edit_message_text(
        "📎 *افزودن عکس/فایل به بارنامه*\n\nشماره بارنامه‌ای که می‌خواهید بهش عکس یا فایل اضافه کنید را وارد کنید:",
        parse_mode="Markdown",
        reply_markup=admin_back_keyboard()
    )


async def admin_stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    db = load_db()
    total = len(db)
    completed = sum(1 for d in db.values() if d.get("status") == "completed")
    pending = total - completed
    total_docs = sum(len(d.get("documents", {})) for d in db.values())

    await query.edit_message_text(
        f"📈 *آمار کلی پی‌بار*\n\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"📦 کل بارنامه‌ها: *{total}*\n"
        f"✅ تکمیل‌شده: *{completed}*\n"
        f"⏳ در انتظار: *{pending}*\n"
        f"📄 کل مستندات: *{total_docs}*\n"
        f"━━━━━━━━━━━━━━━━━━━━",
        parse_mode="Markdown",
        reply_markup=admin_back_keyboard()
    )


async def admin_list_barnames(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    db = load_db()
    if not db:
        await query.edit_message_text(
            "📭 هیچ بارنامه‌ای ثبت نشده.",
            reply_markup=admin_keyboard()
        )
        return

    lines = []
    for barname, data in sorted(db.items(), key=lambda x: x[1].get("created_at", ""), reverse=True):
        status = review_status_label(data)
        doc_count = len(data.get("documents", {}))
        created = to_jalali(data.get("created_at", "-"))
        lines.append(f"{status} | `{barname}` | {doc_count} مستند | {created}")

    text = "📊 *لیست بارنامه‌ها:*\n\n" + "\n".join(lines[:20])
    if len(db) > 20:
        text += f"\n\n... و {len(db)-20} بارنامه دیگر"

    await query.edit_message_text(
        text,
        parse_mode="Markdown",
        reply_markup=admin_back_keyboard()
    )


def build_approved_list_content():
    """متن و کیبورد لیست بارنامه‌های تایید شده در ۵ روز اخیر"""
    db = load_db()
    cutoff = now_tehran() - timedelta(days=5)
    rows = []
    for barname, data in db.items():
        review = data.get("review", {})
        if review.get("status") != "approved":
            continue
        reviewed_at = review.get("reviewed_at", "")
        try:
            reviewed_dt = datetime.strptime(reviewed_at, "%Y-%m-%d %H:%M")
        except (ValueError, TypeError):
            continue
        if reviewed_dt >= cutoff:
            rows.append((barname, data, reviewed_dt))

    if not rows:
        return "📭 در ۵ روز گذشته بارنامه تایید شده‌ای ثبت نشده.", admin_back_keyboard()

    rows.sort(key=lambda x: x[2], reverse=True)
    lines = []
    for barname, data, reviewed_dt in rows:
        lines.append(
            f"✅ `{barname}` | راننده: {data.get('driver_name', '-')} | "
            f"{to_jalali(reviewed_dt.strftime('%Y-%m-%d %H:%M'))}"
        )

    text = "✅ *بارنامه‌های تایید شده (۵ روز اخیر):*\n\n" + "\n".join(lines[:30])
    if len(rows) > 30:
        text += f"\n\n... و {len(rows)-30} بارنامه دیگر"

    return text, admin_back_keyboard()


PENDING_PAGE_SIZE = 8

def build_pending_list_content(page: int = 0):
    """متن و کیبورد لیست بارنامه‌های نیازمند بررسی / تایید نشده (صفحه‌بندی‌شده تا در یک پیام جا بشه)"""
    db = load_db()
    rows = []
    for barname, data in db.items():
        review = data.get("review", {})
        status = review.get("status")
        if status == "approved":
            continue
        if not data.get("documents"):
            continue
        rows.append((barname, data))

    if not rows:
        return "📭 در حال حاضر بارنامه‌ای در انتظار بررسی یا تایید‌نشده وجود ندارد.", admin_back_keyboard()

    rows.sort(key=lambda x: x[1].get("created_at", ""), reverse=True)

    total_pages = max(1, (len(rows) + PENDING_PAGE_SIZE - 1) // PENDING_PAGE_SIZE)
    page = max(0, min(page, total_pages - 1))
    start = page * PENDING_PAGE_SIZE
    page_rows = rows[start:start + PENDING_PAGE_SIZE]

    buttons = []
    for barname, data in page_rows:
        label = f"{review_status_label(data)} — {barname}"
        token = make_barname_token(barname)
        buttons.append([InlineKeyboardButton(label, callback_data=f"admin_open_{token}")])

    nav_row = []
    if page > 0:
        nav_row.append(InlineKeyboardButton("◀️ قبلی", callback_data=f"admin_pending_page_{page-1}"))
    if page < total_pages - 1:
        nav_row.append(InlineKeyboardButton("بعدی ▶️", callback_data=f"admin_pending_page_{page+1}"))
    if nav_row:
        buttons.append(nav_row)

    buttons.append([InlineKeyboardButton("🔙 بازگشت به پنل ادمین", callback_data="admin_back")])

    text = (
        "🕐 *بارنامه‌های نیازمند بررسی / تایید نشده:*\n\n"
        "برای مشاهده‌ی مجدد مستندات و اقدام روی هرکدام، آن را انتخاب کنید 👇\n\n"
        f"_(صفحه {page+1} از {total_pages} — مجموعاً {len(rows)} بارنامه)_"
    )

    return text, InlineKeyboardMarkup(buttons)


async def admin_approved_list(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """لیست بارنامه‌های تایید شده تا ۵ روز گذشته"""
    query = update.callback_query
    await query.answer()
    text, markup = build_approved_list_content()
    await query.edit_message_text(text, parse_mode="Markdown", reply_markup=markup)


async def admin_pending_list(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """لیست بارنامه‌هایی که واکنش نشان داده نشده یا تایید نشده‌اند - قابل انتخاب برای بررسی مجدد"""
    query = update.callback_query
    await query.answer()
    text, markup = build_pending_list_content(page=0)
    await query.edit_message_text(text, parse_mode="Markdown", reply_markup=markup)


async def admin_pending_page(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """جابه‌جایی صفحه در لیست نیازمند بررسی (بدون ارسال پیام جدید)"""
    query = update.callback_query
    await query.answer()
    try:
        page = int(query.data.replace("admin_pending_page_", ""))
    except ValueError:
        page = 0
    text, markup = build_pending_list_content(page=page)
    await query.edit_message_text(text, parse_mode="Markdown", reply_markup=markup)


async def admin_open_barname(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """نمایش مجدد مستندات یک بارنامه‌ی خاص برای ادمین.
    نکته‌ی مهم: این تابع هرگز نباید بارنامه‌ای را که قبلاً «تایید با کسری بار»، «تایید شده»
    یا «عدم تایید» شده دوباره به «در انتظار بررسی» برگرداند — فقط برای بارنامه‌های واقعاً
    جدید/در انتظار، یک پکیج بررسی کامل (با دکمه‌های تایید/رد/کسری) ارسال می‌کند."""
    query = update.callback_query
    user = query.from_user
    if user.id not in ADMIN_IDS:
        logger.warning(f"⛔️ تلاش دسترسی غیرمجاز به دکمه ادمین توسط آیدی {user.id} (لیست ادمین‌های مجاز: {ADMIN_IDS})")
        await query.answer("⛔️ دسترسی ندارید.", show_alert=True)
        return

    token = query.data.replace("admin_open_", "")
    barname = find_barname_by_token(token)
    if not barname:
        await query.answer("⚠️ بارنامه یافت نشد.", show_alert=True)
        return

    db_data = get_barname_data(barname)
    if not db_data.get("documents"):
        await query.answer("⚠️ این بارنامه مستندی ندارد.", show_alert=True)
        return

    review = db_data.get("review", {})
    status = review.get("status")
    driver_link = f"{db_data.get('driver_name', '-')} (آیدی: `{db_data.get('driver_id')}`)"

    if status == "partial":
        # هنوز در انتظار تایید نهاییه — فقط اطلاعات فعلی + دکمه‌ی «تایید نهایی» دوباره فرستاده می‌شود
        await query.answer("در حال ارسال مستندات...")
        for doc_key, doc_info in db_data.get("documents", {}).items():
            try:
                await send_single_doc(
                    context.bot, user.id, doc_key, doc_info, barname,
                    extra_caption=f"👤 فرستنده: {driver_link}", parse_mode="Markdown"
                )
            except Exception as e:
                logger.error(f"خطا در ارسال مستند {doc_key}: {e}")

        resp = review.get("driver_response")
        resp_label = "✅ پذیرفته" if resp == "accepted" else ("❌ نپذیرفته" if resp == "rejected" else "🕐 هنوز پاسخی نداده")
        await _send_partial_finalize_prompt(
            context, barname, db_data, review.get("id", ""),
            text=(
                f"📦 بارنامه {barname} همچنان در وضعیت «تایید با کسری بار» است.\n\n"
                f"⚠️ مقدار کسری: {review.get('deduction_note', '-')}\n"
                f"👤 پاسخ راننده: {resp_label}\n\n"
                "برای تایید نهایی، دکمه‌ی زیر را بزنید 👇"
            )
        )
        return

    if status in ("approved", "rejected"):
        # تصمیم قبلاً نهایی شده — فقط مستندات برای مرور دوباره فرستاده می‌شود، بدون تغییر وضعیت
        await query.answer("در حال ارسال مستندات...")
        for doc_key, doc_info in db_data.get("documents", {}).items():
            try:
                await send_single_doc(
                    context.bot, user.id, doc_key, doc_info, barname,
                    extra_caption=f"👤 فرستنده: {driver_link}", parse_mode="Markdown"
                )
            except Exception as e:
                logger.error(f"خطا در ارسال مستند {doc_key}: {e}")
        await context.bot.send_message(
            chat_id=user.id,
            text=f"📦 بارنامه {barname} — {review_status_label(db_data)}\n(این بارنامه قبلاً نهایی شده و دیگر قابل تغییر نیست.)"
        )
        return

    # status == "pending" یا اصلاً بررسی‌ای شروع نشده → پکیج بررسی کامل با دکمه‌های تایید/رد/کسری
    await query.answer("در حال ارسال مستندات...")
    await dispatch_review_package(context, barname, db_data)


async def admin_text_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user

    # ─── پاسخ راننده به «مورد تایید نیست» برای کسری بار (برای همه کاربران، نه فقط ادمین) ───
    dispute = context.user_data.get("driver_dispute_pending")
    if dispute:
        rid = dispute["rid"]
        barname = dispute["barname"]
        note_text = update.message.text.strip()

        _, db_data = find_barname_by_review_id(rid)
        if not db_data or db_data.get("review", {}).get("status") != "partial" or db_data["review"].get("driver_response"):
            await update.message.reply_text(
                "⚠️ این درخواست دیگر معتبر نیست یا قبلاً پاسخ داده شده.",
                reply_markup=driver_reply_keyboard()
            )
            context.user_data.pop("driver_dispute_pending", None)
            return

        db_data["review"]["driver_response"] = "rejected"
        db_data["review"]["driver_reason"] = note_text
        add_barname_log(
            db_data, "عدم پذیرش کسری بار توسط راننده",
            actor=f"{user.full_name} (راننده)", detail=note_text
        )
        save_barname_data(barname, db_data)

        driver_link = f"{user.full_name} (آیدی: `{user.id}`)"
        for admin_id in ADMIN_IDS:
            try:
                await context.bot.send_message(
                    chat_id=admin_id,
                    text=(
                        f"❌ راننده کسری بار بارنامه شماره {barname} را قبول ندارد.\n\n"
                        f"📝 علت: {note_text}\n"
                        f"👤 راننده: {driver_link}\n\n"
                        "لطفاً جهت هماهنگی، مستقیماً با راننده ارتباط بگیرید."
                    ),
                    parse_mode="Markdown"
                )
            except Exception as e:
                logger.error(f"خطا در اطلاع‌رسانی به ادمین {admin_id}: {e}")

        await _send_partial_finalize_prompt(
            context, barname, db_data, rid,
            text=(
                f"📦 بارنامه شماره {barname} — راننده کسری بار را نپذیرفت.\n\n"
                "پس از هماهنگی با راننده، هر زمان که خواستید می‌توانید با دکمه‌ی زیر "
                "این بارنامه را همچنان تایید نهایی کنید 👇"
            )
        )

        await update.message.reply_text(
            f"راننده محترم با توجه به عدم تایید کسری بار توسط شما، به آیدی ادمین {ADMIN_USERNAME} پیام دهید.",
            reply_markup=driver_reply_keyboard()
        )
        context.user_data.pop("driver_dispute_pending", None)
        return

    if user.id not in ADMIN_IDS:
        return

    # ─── پاسخ ادمین به عدم تایید / تایید با کسری بار ───
    review_pending = context.user_data.get("review_pending")
    if review_pending:
        rid = review_pending["rid"]
        r_action = review_pending["action"]
        note_text = update.message.text.strip()

        barname, db_data = find_barname_by_review_id(rid)
        if not barname or db_data.get("review", {}).get("status") != "pending":
            await update.message.reply_text(
                "⚠️ این درخواست دیگر معتبر نیست یا قبلاً بررسی شده.",
                reply_markup=admin_keyboard()
            )
            context.user_data.pop("review_pending", None)
            return

        db_data["review"]["reviewed_by"] = user.id
        db_data["review"]["reviewed_at"] = now_str()
        driver_id = db_data.get("driver_id")

        if r_action == "reject":
            db_data["review"]["status"] = "rejected"
            db_data["review"]["reason"] = note_text
            add_barname_log(
                db_data, "عدم تایید مستندات",
                actor=f"{user.first_name} (ادمین، از طریق ربات)", detail=note_text
            )
            save_barname_data(barname, db_data)

            if driver_id:
                try:
                    resubmit_token = make_barname_token(barname)
                    await context.bot.send_message(
                        chat_id=driver_id,
                        text=(
                            f"❌ راننده محترم مستندات ارسالی مربوط به بارنامه شماره {barname} "
                            f"به دلیل «{note_text}» مورد تایید قرار نگرفت، "
                            "لطفا مجددا مستندات را به صورت واضح و کامل ارسال نمایید."
                        ),
                        reply_markup=InlineKeyboardMarkup([
                            [InlineKeyboardButton("🔄 بارگذاری مجدد مستندات", callback_data=f"resubmit_{resubmit_token}")]
                        ])
                    )
                    await context.bot.send_message(
                        chat_id=driver_id,
                        text="👇 همچنین می‌توانید از منوی پایین صفحه استفاده کنید.",
                        reply_markup=driver_reply_keyboard()
                    )
                except Exception as e:
                    logger.error(f"خطا در اطلاع‌رسانی به راننده: {e}")

            await update.message.reply_text(
                f"❌ عدم تأیید بارنامه {barname} ثبت و به راننده اطلاع داده شد.",
                reply_markup=admin_keyboard()
            )
            await _finalize_review(context, barname, db_data, f"❌ عدم تأیید (علت: {note_text})")

        elif r_action == "partial":
            db_data["review"]["status"] = "partial"
            db_data["review"]["deduction_note"] = note_text
            db_data["review"]["driver_response"] = None
            add_barname_log(
                db_data, "تایید با کسری بار",
                actor=f"{user.first_name} (ادمین، از طریق ربات)", detail=note_text
            )
            save_barname_data(barname, db_data)

            if driver_id:
                try:
                    await context.bot.send_message(
                        chat_id=driver_id,
                        text=(
                            f"⚠️ راننده محترم طبق بررسی حواله و رسید ارسالی بارنامه شماره {barname}، "
                            f"مقدار {note_text} کسری بار دارید که هزینه آن می‌بایست از مبلغ بارنامه کسر شود.\n\n"
                            "آیا این کسری بار مورد تایید شماست؟"
                        ),
                        reply_markup=driver_partial_response_keyboard(rid)
                    )
                    await context.bot.send_message(
                        chat_id=driver_id,
                        text="👇 همچنین می‌توانید از منوی پایین صفحه استفاده کنید.",
                        reply_markup=driver_reply_keyboard()
                    )
                except Exception as e:
                    logger.error(f"خطا در اطلاع‌رسانی به راننده: {e}")

            await update.message.reply_text(
                f"⚠️ تأیید با کسری بار برای بارنامه {barname} ثبت و به راننده اطلاع داده شد. "
                "پاسخ راننده به‌زودی برای شما ارسال خواهد شد.",
                reply_markup=admin_keyboard()
            )
            await _finalize_review(context, barname, db_data, f"⚠️ تأیید با کسری بار ({note_text}) — در انتظار پاسخ راننده")
            await _send_partial_finalize_prompt(
                context, barname, db_data, rid,
                text=(
                    f"📦 بارنامه {barname} در وضعیت «کسری بار» قرار گرفت.\n\n"
                    "هر زمان که خواستید — چه راننده پاسخ بدهد چه ندهد — می‌توانید با دکمه‌ی زیر "
                    "آن را تایید نهایی کنید 👇"
                )
            )

        context.user_data.pop("review_pending", None)
        return

    action = context.user_data.get("admin_action")
    if not action:
        return

    barname = normalize_barname_digits(update.message.text) or update.message.text.strip()

    # ─── حذف بارنامه ───
    if action == "delete_barname":
        db = load_db()
        if barname not in db:
            await update.message.reply_text(
                f"❌ بارنامه *{barname}* یافت نشد.",
                parse_mode="Markdown",
                reply_markup=admin_keyboard()
            )
        else:
            del db[barname]
            save_db(db)
            await update.message.reply_text(
                f"🗑 بارنامه *{barname}* با موفقیت حذف شد.",
                parse_mode="Markdown",
                reply_markup=admin_keyboard()
            )
        context.user_data.pop("admin_action", None)
        return

    # ─── دریافت شماره بارنامه برای افزودن عکس/فایل ───
    if action == "attach_barname":
        db_data = get_barname_data(barname)
        if not db_data.get("created_at"):
            await update.message.reply_text(
                f"❌ بارنامه *{barname}* یافت نشد. لطفاً دوباره شماره بارنامه را وارد کنید یا /start را بزنید.",
                parse_mode="Markdown",
                reply_markup=admin_keyboard()
            )
            context.user_data.pop("admin_action", None)
            return

        context.user_data["admin_action"] = "attach_file"
        context.user_data["admin_attach_barname"] = barname
        await update.message.reply_text(
            f"📎 حالا عکس یا فایل مورد نظر برای بارنامه *{barname}* را ارسال کنید.\n"
            "_(می‌توانید یک توضیح کوتاه هم به‌عنوان کپشن همراه عکس/فایل بنویسید — اختیاری)_",
            parse_mode="Markdown"
        )
        return

    # ─── دریافت مستندات ───
    if action == "get_docs":
        db_data = get_barname_data(barname)

        if not db_data.get("documents"):
            await update.message.reply_text(
                f"❌ بارنامه *{barname}* یافت نشد یا مستندی ندارد.",
                parse_mode="Markdown",
                reply_markup=admin_keyboard()
            )
            context.user_data.pop("admin_action", None)
            return

        docs = db_data["documents"]
        driver_id = db_data.get("driver_id")
        driver_link = f"{db_data.get('driver_name', '-')} (آیدی: `{driver_id}`)" if driver_id else db_data.get('driver_name', '-')
        await update.message.reply_text(
            f"📦 *بارنامه {barname}*\n"
            f"👤 راننده: {driver_link}\n"
            f"📅 تاریخ: {to_jalali(db_data.get('created_at', '-'))}\n"
            f"📄 تعداد: {len(docs)} مستند\n"
            f"📌 وضعیت: {review_status_label(db_data)}\n\n"
            "در حال ارسال...",
            parse_mode="Markdown"
        )

        for doc_key, doc_info in docs.items():
            try:
                await send_single_doc(
                    context.bot, user.id, doc_key, doc_info, barname,
                    extra_caption=f"👤 فرستنده: {driver_link}\n🕐 {to_jalali(doc_info.get('uploaded_at', '-'))}",
                    parse_mode="Markdown"
                )
            except Exception as e:
                label = doc_label(doc_key, doc_info)
                await update.message.reply_text(f"⚠️ خطا در ارسال {label}: {e}")

        await update.message.reply_text(
            f"✅ همه مستندات بارنامه *{barname}* ارسال شد.",
            parse_mode="Markdown",
            reply_markup=admin_keyboard()
        )
        context.user_data.pop("admin_action", None)


async def admin_attach_receive_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """دریافت عکس/فایلی که ادمین می‌خواهد به یک بارنامه اضافه کند (خارج از فلوی مکالمه‌ی راننده)"""
    user = update.effective_user
    if user.id not in ADMIN_IDS:
        return
    if context.user_data.get("admin_action") != "attach_file":
        return

    barname = context.user_data.get("admin_attach_barname")
    if not barname:
        context.user_data.pop("admin_action", None)
        return

    photo = update.message.photo[-1] if update.message.photo else None
    doc = update.message.document
    caption = (update.message.caption or "").strip()

    if photo:
        file_id, file_type = photo.file_id, "photo"
    elif doc:
        mime = (doc.mime_type or "").lower()
        fname = (doc.file_name or "").lower()
        is_pdf = (mime == "application/pdf") or fname.endswith(".pdf")
        if not is_pdf:
            await update.message.reply_text(
                "⚠️ فقط *عکس* یا فایل *PDF* قابل قبول است.\nلطفاً به‌صورت عکس یا PDF ارسال کنید.",
                parse_mode="Markdown"
            )
            return
        file_id, file_type = doc.file_id, "pdf"
    else:
        await update.message.reply_text("⚠️ لطفاً یک عکس یا فایل PDF ارسال کنید.")
        return

    data = get_barname_data(barname)
    if not data.get("created_at"):
        await update.message.reply_text(
            f"❌ بارنامه *{barname}* دیگر پیدا نشد (شاید حذف شده).",
            parse_mode="Markdown",
            reply_markup=admin_keyboard()
        )
        context.user_data.pop("admin_action", None)
        context.user_data.pop("admin_attach_barname", None)
        return

    attachments = data.setdefault("admin_attachments", [])
    attachments.append({
        "id": uuid.uuid4().hex[:10],
        "source": "bot",
        "file_id": file_id,
        "file_type": file_type,
        "caption": caption,
        "uploaded_by": f"{user.first_name} (ادمین)",
        "uploaded_at": now_str(),
    })
    add_barname_log(
        data, "افزودن عکس/فایل توسط ادمین",
        actor=f"{user.first_name} (ادمین، از طریق ربات)", detail=caption
    )
    save_barname_data(barname, data)

    context.user_data.pop("admin_action", None)
    context.user_data.pop("admin_attach_barname", None)

    await update.message.reply_text(
        f"✅ عکس/فایل به بارنامه *{barname}* اضافه شد و در داشبورد هم قابل مشاهده است.\n\n"
        "برای افزودن مورد دیگر به همین بارنامه یا بارنامه‌ی دیگر، دوباره از منو «📎 افزودن عکس/فایل» را بزنید.",
        parse_mode="Markdown",
        reply_markup=admin_keyboard()
    )


async def reply_keyboard_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """هندلر دکمه‌های منوی پایین صفحه"""
    text = update.message.text
    user = update.effective_user
    is_admin = user.id in ADMIN_IDS

    # ─── دکمه‌های راننده ───
    if text == "🚀 بارگزاری مدارک":
        await update.message.reply_text(
            "📝 لطفاً *شماره بارنامه* را وارد کنید:",
            parse_mode="Markdown",
            reply_markup=barname_entry_keyboard()
        )
        return WAIT_BARNAME

    elif text == "📋 راهنما":
        await update.message.reply_text(
            "📋 *راهنمای استفاده از ربات پی‌بار*\n\n"
            "━━━━━━━━━━━━━━━━━━━━\n"
            "🔹 *مستندات مورد نیاز:*\n"
            "۱. 📄 اصل بارنامه\n"
            "۲. 🔵 حواله خروج مبدأ (محل بارگیری)\n"
            "۳. 🔴 رسید تخلیه مقصد\n"
            "۴. 💳 شماره حساب بانک تجارت (جهت پرداخت سریع‌تر) یا شماره شبا، به نام راننده یا شرکت باربری\n"
            "   (در صورت ارسال شماره حساب شخص دیگر، رضایت‌نامه به همراه شماره ملی صاحب حساب هم ارسال شود)\n\n"
            "━━━━━━━━━━━━━━━━━━━━\n"
            "💡 *نکات مهم:*\n"
            "• عکس‌ها باید واضح و خوانا باشند\n"
            "• لازم نیست به ترتیب خاصی ارسال کنید، هرکدام را که آماده دارید بفرستید\n"
            "• وقتی همه‌ی مدارک را فرستادید، دکمه‌ی «✅ تایید نهایی...» پایین صفحه را بزنید\n"
            "• /cancel برای لغو کامل عملیات",
            parse_mode="Markdown",
            reply_markup=driver_reply_keyboard()
        )

    elif text == "📦 وضعیت بارنامه":
        rows = get_driver_barnames(user.id)

        if not rows:
            await update.message.reply_text(
                "📭 شما تا الان هیچ بارنامه‌ای ثبت نکرده‌اید.\n"
                "برای شروع دکمه 🚀 بارگزاری مدارک را بزنید.",
                reply_markup=driver_reply_keyboard()
            )
        else:
            lines = []
            for bn, data in rows[:20]:
                doc_count = len(data.get("documents", {}))
                status = review_status_label(data)
                lines.append(f"{status} | `{bn}` | {doc_count} مدرک | {to_jalali(data.get('created_at', '-'))}")

            text_msg = "📦 *وضعیت بارنامه‌های شما:*\n\n" + "\n".join(lines)
            if len(rows) > 20:
                text_msg += f"\n\n... و {len(rows)-20} بارنامه دیگر"

            # اگه راننده در یک بارنامه‌ی رد شده گیر کرده، دکمه‌ی بارگذاری مجدد رو هم نشون بده
            last_rejected = next((bn for bn, d in rows if d.get("review", {}).get("status") == "rejected"), None)
            markup = None
            if last_rejected:
                token = make_barname_token(last_rejected)
                markup = InlineKeyboardMarkup([
                    [InlineKeyboardButton(f"🔄 بارگذاری مجدد {last_rejected}", callback_data=f"resubmit_{token}")]
                ])

            await update.message.reply_text(text_msg, parse_mode="Markdown", reply_markup=driver_reply_keyboard())
            if markup:
                await update.message.reply_text("برای بارگذاری مجدد مستندات رد شده 👇", reply_markup=markup)

    elif text == "❌ لغو عملیات":
        context.user_data.clear()
        await update.message.reply_text(
            "❌ عملیات لغو شد.\nهر زمان آماده بودید دوباره شروع کنید.",
            reply_markup=driver_reply_keyboard()
        )
        return ConversationHandler.END

    # ─── دکمه‌های ادمین ───
    elif is_admin and text == "🔍 دریافت مستندات":
        context.user_data["admin_action"] = "get_docs"
        await update.message.reply_text(
            "🔍 شماره بارنامه مورد نظر را وارد کنید:",
            reply_markup=admin_reply_keyboard()
        )

    elif is_admin and text == "📊 لیست بارنامه‌ها":
        db = load_db()
        if not db:
            await update.message.reply_text("📭 هیچ بارنامه‌ای ثبت نشده.", reply_markup=admin_reply_keyboard())
            return
        lines = []
        for bn, data in sorted(db.items(), key=lambda x: x[1].get("created_at",""), reverse=True):
            status = review_status_label(data)
            lines.append(f"{status} | `{bn}` | {len(data.get('documents',{}))} مستند | {to_jalali(data.get('created_at','-'))}")
        text_msg = "📊 *لیست بارنامه‌ها:*\n\n" + "\n".join(lines[:20])
        if len(db) > 20:
            text_msg += f"\n\n... و {len(db)-20} بارنامه دیگر"
        await update.message.reply_text(text_msg, parse_mode="Markdown", reply_markup=admin_reply_keyboard())

    elif is_admin and text == "✅ بارنامه‌های تایید شده":
        list_text, _ = build_approved_list_content()
        await update.message.reply_text(list_text, parse_mode="Markdown", reply_markup=admin_reply_keyboard())

    elif is_admin and text == "🕐 نیازمند بررسی":
        list_text, list_markup = build_pending_list_content()
        await update.message.reply_text(list_text, parse_mode="Markdown", reply_markup=list_markup)

    elif is_admin and text == "📈 آمار کلی":
        db = load_db()
        total = len(db)
        completed = sum(1 for d in db.values() if d.get("status") == "completed")
        total_docs = sum(len(d.get("documents", {})) for d in db.values())
        await update.message.reply_text(
            f"📈 *آمار کلی پی‌بار*\n\n"
            f"📦 کل بارنامه‌ها: *{total}*\n"
            f"✅ تکمیل‌شده: *{completed}*\n"
            f"⏳ در انتظار: *{total - completed}*\n"
            f"📄 کل مستندات: *{total_docs}*",
            parse_mode="Markdown",
            reply_markup=admin_reply_keyboard()
        )

    elif is_admin and text == "🗑 حذف بارنامه":
        context.user_data["admin_action"] = "delete_barname"
        await update.message.reply_text(
            "🗑 شماره بارنامه‌ای که می‌خواهید حذف کنید را وارد کنید:",
            reply_markup=admin_reply_keyboard()
        )

    elif is_admin and text == "📎 افزودن عکس/فایل":
        context.user_data["admin_action"] = "attach_barname"
        context.user_data.pop("admin_attach_barname", None)
        await update.message.reply_text(
            "📎 شماره بارنامه‌ای که می‌خواهید بهش عکس یا فایل اضافه کنید را وارد کنید:",
            reply_markup=admin_reply_keyboard()
        )

    elif is_admin and text == "🏠 منوی اصلی ادمین":
        context.user_data.clear()
        await update.message.reply_text(
            "🔐 *پنل مدیریت پی‌بار*\nاز منوی پایین یا دکمه‌های زیر انتخاب کنید:",
            parse_mode="Markdown",
            reply_markup=admin_reply_keyboard()
        )
        await update.message.reply_text("👇", reply_markup=admin_keyboard())

    return ConversationHandler.END


async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    _cancel_upload_reminder(context, update.effective_chat.id)
    context.user_data.clear()
    user = update.effective_user
    if user.id in ADMIN_IDS:
        await update.message.reply_text(
            "❌ عملیات لغو شد.\n\n🔐 *پنل مدیریت پی‌بار*",
            parse_mode="Markdown",
            reply_markup=admin_reply_keyboard()
        )
    else:
        await update.message.reply_text(
            "❌ عملیات لغو شد.\n\nبرای شروع مجدد از منوی پایین صفحه استفاده کنید:",
            parse_mode="Markdown",
            reply_markup=driver_reply_keyboard()
        )
    return ConversationHandler.END


# ─────────── راه‌اندازی ───────────

# ─────────── داشبورد وب ادمین (FastAPI) ───────────

dashboard_api = FastAPI(title="داشبورد پی‌بار")


def _resolve_dashboard_role(token: str) -> str:
    """توکن را بررسی می‌کند و نقش متناظر را برمی‌گرداند: 'admin' یا 'viewer'.
    برای درخواست‌های نامعتبر، خطای مناسب را raise می‌کند."""
    if not DASHBOARD_TOKEN:
        raise HTTPException(status_code=503, detail="DASHBOARD_TOKEN روی سرور تنظیم نشده است.")
    if token and token == DASHBOARD_TOKEN:
        return "admin"
    if token and DASHBOARD_VIEWER_TOKEN and token == DASHBOARD_VIEWER_TOKEN:
        return "viewer"
    raise HTTPException(status_code=401, detail="توکن نامعتبر است.")


def _check_dashboard_token(token: str) -> str:
    """هر توکن معتبری (ادمین یا مشاهده‌گر) را قبول می‌کند — برای endpointهای فقط-خواندنی."""
    return _resolve_dashboard_role(token)


def _require_admin_token(token: str) -> str:
    """فقط توکن ادمین را قبول می‌کند — برای endpointهایی که داده را تغییر می‌دهند
    (تایید/رد/کسری/پیام/بارگزاری/نوع محصول). توکن مشاهده‌گر اینجا رد می‌شود."""
    role = _resolve_dashboard_role(token)
    if role != "admin":
        raise HTTPException(status_code=403, detail="این عملیات فقط برای ادمین مجاز است — شما فقط دسترسی مشاهده دارید.")
    return role


@dashboard_api.get("/api/whoami")
async def api_whoami(x_dashboard_token: str = Header(default="")):
    """نقش صاحب توکن فعلی را برمی‌گرداند تا داشبورد بداند دکمه‌های اقدام را نشان بدهد یا نه."""
    role = _resolve_dashboard_role(x_dashboard_token)
    return {"role": role}


@dashboard_api.get("/", response_class=HTMLResponse)
async def dashboard_index():
    html_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "dashboard.html")
    try:
        with open(html_path, encoding="utf-8") as f:
            return HTMLResponse(f.read())
    except FileNotFoundError:
        return HTMLResponse("<h3>فایل dashboard.html پیدا نشد.</h3>", status_code=500)


@dashboard_api.get("/api/barnames")
async def api_list_barnames(
    x_dashboard_token: str = Header(default=""),
    status: str = Query(default="all"),
    search: str = Query(default=""),
):
    _check_dashboard_token(x_dashboard_token)
    db = load_db()
    rows = []
    for bn, data in db.items():
        if search and search.strip() not in bn:
            continue
        review = data.get("review", {})
        st = review.get("status") or "none"
        if status != "all" and st != status:
            continue
        first_resp, last_resp = admin_response_times(data)
        rows.append({
            "barname": bn,
            "driver_name": data.get("driver_name", "-"),
            "driver_id": data.get("driver_id"),
            "status": st,
            "status_label": review_status_label(data),
            "doc_count": len(data.get("documents", {})),
            "created_at": data.get("created_at", "-"),
            "reviewed_at": review.get("reviewed_at"),
            "deduction_note": review.get("deduction_note", ""),
            "product_type": data.get("product_type", ""),
            "first_admin_response_at": first_resp,
            "last_admin_response_at": last_resp,
        })
    rows.sort(key=lambda r: r["created_at"] or "", reverse=True)
    return rows


@dashboard_api.get("/api/export/xlsx")
async def api_export_xlsx(token: str = Query(default="")):
    """خروجی اکسل کامل از همه‌ی بارنامه‌ها — برای گزارش‌گیری"""
    _check_dashboard_token(token)
    db = load_db()

    wb = Workbook()
    ws = wb.active
    ws.title = "بارنامه‌ها"
    ws.sheet_view.rightToLeft = True

    headers = [
        "شماره بارنامه", "نام راننده", "آیدی راننده", "نوع محصول", "وضعیت",
        "یادداشت کسری بار", "پاسخ راننده به کسری", "تعداد مدارک",
        "تاریخ ثبت", "تاریخ ارسال نهایی", "اولین واکنش ادمین", "آخرین واکنش ادمین",
    ]
    ws.append(headers)

    driver_response_labels = {"accepted": "پذیرفته", "rejected": "نپذیرفته"}

    rows = sorted(db.items(), key=lambda x: x[1].get("created_at", "") or "", reverse=True)
    for barname, data in rows:
        review = data.get("review", {})
        first_resp, last_resp = admin_response_times(data)
        dr = review.get("driver_response")
        dr_label = driver_response_labels.get(dr, "-")
        ws.append([
            barname,
            data.get("driver_name", "-"),
            data.get("driver_id", "-"),
            data.get("product_type", "") or "-",
            review_status_label(data),
            review.get("deduction_note", "") or "-",
            dr_label,
            len(data.get("documents", {})),
            data.get("created_at", "-") or "-",
            data.get("completed_at", "-") or "-",
            first_resp or "-",
            last_resp or "-",
        ])

    widths = [16, 20, 14, 18, 26, 22, 16, 10, 16, 16, 16, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"pibar-report-{now_tehran().strftime('%Y%m%d-%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@dashboard_api.get("/api/barnames/{barname}")
async def api_get_barname(barname: str, x_dashboard_token: str = Header(default="")):
    _check_dashboard_token(x_dashboard_token)
    data = get_barname_data(barname)
    if not data.get("documents"):
        raise HTTPException(status_code=404, detail="بارنامه یافت نشد یا مستندی ندارد.")

    docs = []
    for key, info in data.get("documents", {}).items():
        docs.append({
            "key": key,
            "label": doc_label(key, info),
            "file_type": info.get("file_type"),
            "text": info.get("text", ""),
            "uploaded_at": info.get("uploaded_at", ""),
        })

    first_resp, last_resp = admin_response_times(data)
    return {
        "barname": barname,
        "driver_name": data.get("driver_name", "-"),
        "driver_id": data.get("driver_id"),
        "created_at": data.get("created_at"),
        "completed_at": data.get("completed_at"),
        "review": data.get("review", {}),
        "status_label": review_status_label(data),
        "documents": docs,
        "admin_attachments": data.get("admin_attachments", []),
        "product_type": data.get("product_type", ""),
        "first_admin_response_at": first_resp,
        "last_admin_response_at": last_resp,
        "log": data.get("log", []),
    }


@dashboard_api.post("/api/barnames/{barname}/product-type")
async def api_set_product_type(barname: str, x_dashboard_token: str = Header(default=""), body: dict = Body(...)):
    """ثبت/ویرایش نوع محصول یک بارنامه (مثلاً برنج ایرانی، برنج خارجی، شکر، آرد و ...)"""
    _require_admin_token(x_dashboard_token)
    product_type = (body.get("type") or "").strip()

    data = get_barname_data(barname)
    if not data.get("created_at"):
        raise HTTPException(status_code=404, detail="بارنامه یافت نشد.")

    old_value = (data.get("product_type") or "").strip()
    data["product_type"] = product_type
    if product_type != old_value:
        add_barname_log(
            data, "ثبت/ویرایش نوع محصول",
            actor="ادمین (از طریق داشبورد وب)",
            detail=product_type or "(پاک شد)"
        )
    save_barname_data(barname, data)
    return {"ok": True, "product_type": product_type}


@dashboard_api.get("/api/barnames/{barname}/log")
async def api_get_barname_log(barname: str, x_dashboard_token: str = Header(default="")):
    """تاریخچه‌ی کامل تغییرات یک بارنامه (برای نمایش لاگ در داشبورد)"""
    _check_dashboard_token(x_dashboard_token)
    data = get_barname_data(barname)
    if not data.get("created_at"):
        raise HTTPException(status_code=404, detail="بارنامه یافت نشد.")
    log = data.get("log", [])
    # جدیدترین رویداد اول نمایش داده شود
    log_sorted = list(reversed(log))
    return {"barname": barname, "count": len(log_sorted), "log": log_sorted}


def _media_type_for(file_type: str) -> str:
    """نوع MIME مناسب برای هر نوع مستند — تا مرورگر/داشبورد بتواند فرمت را درست تشخیص بدهد"""
    if file_type == "photo":
        return "image/jpeg"
    if file_type == "pdf":
        return "application/pdf"
    return "application/octet-stream"


@dashboard_api.get("/api/barnames/{barname}/doc/{doc_key}/file")
async def api_get_doc_file(barname: str, doc_key: str, token: str = Query(default="")):
    _check_dashboard_token(token)
    data = get_barname_data(barname)
    doc = data.get("documents", {}).get(doc_key)
    if not doc or not doc.get("file_id"):
        raise HTTPException(status_code=404, detail="فایل یافت نشد.")
    if not BOT_INSTANCE:
        raise HTTPException(status_code=503, detail="ربات هنوز آماده نیست.")
    try:
        tg_file = await BOT_INSTANCE.get_file(doc["file_id"])
        file_bytes = await tg_file.download_as_bytearray()
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"خطا در دریافت فایل از بله: {e}")
    media_type = _media_type_for(doc.get("file_type"))
    return StreamingResponse(io.BytesIO(bytes(file_bytes)), media_type=media_type)


@dashboard_api.get("/api/barnames/{barname}/attachment/{attachment_id}/file")
async def api_get_attachment_file(barname: str, attachment_id: str, token: str = Query(default="")):
    _check_dashboard_token(token)
    data = get_barname_data(barname)
    att = next((a for a in data.get("admin_attachments", []) if a.get("id") == attachment_id), None)
    if not att:
        raise HTTPException(status_code=404, detail="پیوست یافت نشد.")

    media_type = _media_type_for(att.get("file_type"))

    if att.get("source") == "dashboard":
        file_path = os.path.join(DATA_DIR, "attachments", barname, att.get("filename", ""))
        if not os.path.isfile(file_path):
            raise HTTPException(status_code=404, detail="فایل روی سرور پیدا نشد.")
        with open(file_path, "rb") as f:
            file_bytes = f.read()
        return StreamingResponse(io.BytesIO(file_bytes), media_type=media_type)

    # source == "bot" — از طریق فایل‌سرور بله دانلود می‌شود
    if not att.get("file_id"):
        raise HTTPException(status_code=404, detail="فایل یافت نشد.")
    if not BOT_INSTANCE:
        raise HTTPException(status_code=503, detail="ربات هنوز آماده نیست.")
    try:
        tg_file = await BOT_INSTANCE.get_file(att["file_id"])
        file_bytes = await tg_file.download_as_bytearray()
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"خطا در دریافت فایل از بله: {e}")
    return StreamingResponse(io.BytesIO(bytes(file_bytes)), media_type=media_type)


@dashboard_api.post("/api/barnames/{barname}/attachment")
async def api_upload_attachment(
    barname: str,
    x_dashboard_token: str = Header(default=""),
    file: UploadFile = File(...),
    caption: str = Form(default=""),
):
    """بارگزاری مستقیم عکس/فایل برای یک بارنامه از داخل داشبورد (بدون نیاز به چت با ربات)"""
    _require_admin_token(x_dashboard_token)
    data = get_barname_data(barname)
    if not data.get("created_at"):
        raise HTTPException(status_code=404, detail="بارنامه یافت نشد.")

    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="فایل خالی است.")
    if len(raw) > 20 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="حجم فایل نباید بیشتر از ۲۰ مگابایت باشد.")

    content_type = (file.content_type or "").lower()
    fname_lower = (file.filename or "").lower()
    if content_type.startswith("image/"):
        file_type = "photo"
    elif content_type == "application/pdf" or fname_lower.endswith(".pdf"):
        file_type = "pdf"
    else:
        raise HTTPException(status_code=400, detail="فقط فایل تصویری یا PDF قابل قبول است.")

    att_id = uuid.uuid4().hex[:10]
    ext = os.path.splitext(file.filename or "")[1][:10]
    stored_name = f"{att_id}{ext}"
    dir_path = os.path.join(DATA_DIR, "attachments", barname)
    os.makedirs(dir_path, exist_ok=True)
    with open(os.path.join(dir_path, stored_name), "wb") as f:
        f.write(raw)

    attachments = data.setdefault("admin_attachments", [])
    attachments.append({
        "id": att_id,
        "source": "dashboard",
        "filename": stored_name,
        "file_type": file_type,
        "caption": (caption or "").strip(),
        "uploaded_by": "ادمین (از طریق داشبورد وب)",
        "uploaded_at": now_str(),
    })
    add_barname_log(
        data, "افزودن عکس/فایل توسط ادمین",
        actor="ادمین (از طریق داشبورد وب)", detail=(caption or "").strip()
    )
    save_barname_data(barname, data)
    return {"ok": True, "id": att_id}


@dashboard_api.post("/api/barnames/{barname}/approve")
async def api_approve(barname: str, x_dashboard_token: str = Header(default="")):
    _require_admin_token(x_dashboard_token)
    data = get_barname_data(barname)
    review = data.get("review", {})
    was_partial = review.get("status") == "partial"
    if review.get("status") not in ("pending", "partial"):
        raise HTTPException(status_code=400, detail="این بارنامه در وضعیتی نیست که بتوان تایید نهایی کرد.")

    review["status"] = "approved"
    review["reviewed_by"] = "dashboard"
    review["reviewed_at"] = now_str()
    data["review"] = review
    if was_partial:
        resp = review.get("driver_response")
        resp_label = "پذیرفته بود" if resp == "accepted" else ("نپذیرفته بود" if resp == "rejected" else "هنوز پاسخی نداده بود")
        add_barname_log(
            data, "تایید نهایی بارنامه (پس از کسری بار)",
            actor="ادمین (از طریق داشبورد وب)", detail=f"راننده کسری بار را {resp_label}"
        )
    else:
        add_barname_log(data, "تایید مستندات", actor="ادمین (از طریق داشبورد وب)")
    save_barname_data(barname, data)

    driver_id = data.get("driver_id")
    if driver_id and BOT_INSTANCE:
        try:
            msg_text = (
                f"✅ راننده محترم بارنامه شماره {barname} شما به‌صورت نهایی تایید شد "
                "و ظرف ۲ روز کاری هزینه بارنامه (با احتساب کسری بار توافق‌شده) به شماره حساب اعلامی شما واریز خواهد شد."
            ) if was_partial else (
                f"✅ راننده محترم مستندات ارسالی بارنامه شماره {barname} شما تایید شد "
                "و ظرف ۲ روز کاری هزینه بارنامه به شماره حساب اعلامی شما واریز خواهد شد."
            )
            await BOT_INSTANCE.send_message(chat_id=driver_id, text=msg_text, reply_markup=driver_reply_keyboard())
        except Exception as e:
            logger.error(f"خطا در اطلاع‌رسانی به راننده از داشبورد: {e}")

    fake_ctx = SimpleNamespace(bot=BOT_INSTANCE)
    if was_partial:
        finalize_messages = data.get("review", {}).get("finalize_messages", {})
        for admin_id_str, msg_id in finalize_messages.items():
            try:
                await BOT_INSTANCE.edit_message_reply_markup(chat_id=int(admin_id_str), message_id=msg_id, reply_markup=None)
            except Exception as e:
                logger.error(f"خطا در حذف دکمه تایید نهایی ادمین {admin_id_str}: {e}")
        await _finalize_review(fake_ctx, barname, data, "✅ تایید نهایی شد (پس از کسری بار — از طریق داشبورد وب)")
    else:
        await _finalize_review(fake_ctx, barname, data, "✅ تأیید شد (از طریق داشبورد وب)")
    return {"ok": True}


@dashboard_api.post("/api/barnames/{barname}/reject")
async def api_reject(barname: str, payload: dict = Body(...), x_dashboard_token: str = Header(default="")):
    _require_admin_token(x_dashboard_token)
    reason = (payload.get("reason") or "").strip()
    if not reason:
        raise HTTPException(status_code=400, detail="علت عدم تایید الزامی است.")

    data = get_barname_data(barname)
    review = data.get("review", {})
    if review.get("status") != "pending":
        raise HTTPException(status_code=400, detail="این بارنامه در وضعیت «در انتظار بررسی» نیست.")

    review["status"] = "rejected"
    review["reason"] = reason
    review["reviewed_by"] = "dashboard"
    review["reviewed_at"] = now_str()
    data["review"] = review
    add_barname_log(data, "عدم تایید مستندات", actor="ادمین (از طریق داشبورد وب)", detail=reason)
    save_barname_data(barname, data)

    driver_id = data.get("driver_id")
    if driver_id and BOT_INSTANCE:
        try:
            resubmit_token = make_barname_token(barname)
            await BOT_INSTANCE.send_message(
                chat_id=driver_id,
                text=(
                    f"❌ راننده محترم مستندات ارسالی مربوط به بارنامه شماره {barname} "
                    f"به دلیل «{reason}» مورد تایید قرار نگرفت، "
                    "لطفا مجددا مستندات را به صورت واضح و کامل ارسال نمایید."
                ),
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("🔄 بارگذاری مجدد مستندات", callback_data=f"resubmit_{resubmit_token}")]
                ])
            )
            await BOT_INSTANCE.send_message(
                chat_id=driver_id,
                text="👇 همچنین می‌توانید از منوی پایین صفحه استفاده کنید.",
                reply_markup=driver_reply_keyboard()
            )
        except Exception as e:
            logger.error(f"خطا در اطلاع‌رسانی به راننده از داشبورد: {e}")

    fake_ctx = SimpleNamespace(bot=BOT_INSTANCE)
    await _finalize_review(fake_ctx, barname, data, f"❌ عدم تأیید (علت: {reason}) — از طریق داشبورد وب")
    return {"ok": True}


@dashboard_api.post("/api/barnames/{barname}/partial")
async def api_partial(barname: str, payload: dict = Body(...), x_dashboard_token: str = Header(default="")):
    _require_admin_token(x_dashboard_token)
    note = (payload.get("note") or "").strip()
    if not note:
        raise HTTPException(status_code=400, detail="مقدار کسری بار الزامی است.")

    data = get_barname_data(barname)
    review = data.get("review", {})
    if review.get("status") != "pending":
        raise HTTPException(status_code=400, detail="این بارنامه در وضعیت «در انتظار بررسی» نیست.")

    rid = review.get("id", "")
    review["status"] = "partial"
    review["deduction_note"] = note
    review["driver_response"] = None
    review["reviewed_by"] = "dashboard"
    review["reviewed_at"] = now_str()
    data["review"] = review
    add_barname_log(data, "تایید با کسری بار", actor="ادمین (از طریق داشبورد وب)", detail=note)
    save_barname_data(barname, data)

    driver_id = data.get("driver_id")
    if driver_id and BOT_INSTANCE:
        try:
            await BOT_INSTANCE.send_message(
                chat_id=driver_id,
                text=(
                    f"⚠️ راننده محترم طبق بررسی حواله و رسید ارسالی بارنامه شماره {barname}، "
                    f"مقدار {note} کسری بار دارید که هزینه آن می‌بایست از مبلغ بارنامه کسر شود.\n\n"
                    "آیا این کسری بار مورد تایید شماست؟"
                ),
                reply_markup=driver_partial_response_keyboard(rid)
            )
            await BOT_INSTANCE.send_message(
                chat_id=driver_id,
                text="👇 همچنین می‌توانید از منوی پایین صفحه استفاده کنید.",
                reply_markup=driver_reply_keyboard()
            )
        except Exception as e:
            logger.error(f"خطا در اطلاع‌رسانی به راننده از داشبورد: {e}")

    fake_ctx = SimpleNamespace(bot=BOT_INSTANCE)
    await _finalize_review(
        fake_ctx, barname, data,
        f"⚠️ تأیید با کسری بار ({note}) — از طریق داشبورد وب — در انتظار پاسخ راننده"
    )
    return {"ok": True}


@dashboard_api.post("/api/barnames/{barname}/message")
async def api_custom_message(barname: str, payload: dict = Body(...), x_dashboard_token: str = Header(default="")):
    _require_admin_token(x_dashboard_token)
    text = (payload.get("text") or "").strip()
    if not text:
        raise HTTPException(status_code=400, detail="متن پیام الزامی است.")

    data = get_barname_data(barname)
    driver_id = data.get("driver_id")
    if not driver_id:
        raise HTTPException(status_code=404, detail="راننده‌ای برای این بارنامه یافت نشد.")
    if not BOT_INSTANCE:
        raise HTTPException(status_code=503, detail="ربات هنوز آماده نیست.")

    try:
        await BOT_INSTANCE.send_message(chat_id=driver_id, text=text)
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"خطا در ارسال پیام: {e}")

    add_barname_log(data, "پیام دستی به راننده", actor="ادمین (از طریق داشبورد وب)", detail=text)
    save_barname_data(barname, data)

    return {"ok": True}


async def _on_startup(application):
    """قبل از شروع polling، هر webhook فعال روی این ربات را حذف می‌کند تا هرگز polling و webhook هم‌زمان فعال نباشند
    (فعال بودن هم‌زمان هر دو، دلیل رایج ارسال دوبار هر پیام است)."""
    try:
        info = await application.bot.get_webhook_info()
        if info and info.url:
            logger.warning(f"⚠️ یک webhook فعال روی این ربات پیدا شد ({info.url}) — در حال حذف آن تا فقط polling فعال باشد...")
        deleted = await application.bot.delete_webhook(drop_pending_updates=True)
        logger.info(f"✅ بررسی/حذف webhook انجام شد (نتیجه: {deleted}).")
    except Exception as e:
        logger.error(f"⚠️ خطا در بررسی/حذف webhook: {e}")


async def run_app():
    global BOT_INSTANCE

    # شناسه‌ی یکتای این نمونه از پروسه — اگر در لاگ‌ها دو INSTANCE_ID متفاوت هم‌زمان دیدید،
    # یعنی دو نسخه از ربات هم‌زمان در حال اجرا هستند و همین باعث ارسال دوبار هر پیام می‌شود.
    instance_id = uuid.uuid4().hex[:8]
    logger.info(f"🆔 INSTANCE_ID این اجرا: {instance_id} — اگر جای دیگری هم‌زمان یک INSTANCE_ID دیگر دیدید، دو نمونه از ربات هم‌زمان اجرا شده‌اند.")

    # تشخیص مشکلات رایج پیکربندی قبل از شروع
    if not ADMIN_IDS:
        logger.warning("⚠️ متغیر ADMIN_IDS خالی است — هیچ ادمینی به پنل مدیریت دسترسی نخواهد داشت!")
    else:
        logger.info(f"✅ آیدی ادمین‌های تعریف‌شده: {ADMIN_IDS}")

    if not DASHBOARD_TOKEN:
        logger.warning("⚠️ متغیر DASHBOARD_TOKEN تنظیم نشده — داشبورد وب غیرفعال خواهد بود (همه‌ی درخواست‌ها 503 می‌گیرند).")
    elif DASHBOARD_VIEWER_TOKEN:
        logger.info("👁️ نقش «مشاهده‌گر» برای داشبورد فعال است (DASHBOARD_VIEWER_TOKEN تنظیم شده).")

    if os.path.exists(DB_FILE):
        try:
            existing = load_db()
            logger.info(f"📦 دیتابیس موجود بارگذاری شد از «{DB_FILE}» — {len(existing)} بارنامه ثبت‌شده.")
        except Exception as e:
            logger.error(f"⚠️ خطا در خواندن دیتابیس موجود: {e}")
    else:
        logger.warning(
            f"⚠️ فایل دیتابیس («{DB_FILE}») پیدا نشد — یک دیتابیس خالی جدید ساخته می‌شود. "
            "اگر این پیام را بعد از هر ری‌استارت می‌بینید، یعنی فضای ذخیره‌سازی سرور شما پایدار (persistent) نیست "
            "و تمام بارنامه‌ها/تاریخچه‌ها با هر دیپلوی یا ری‌استارت پاک می‌شوند — "
            "برای رفع این مشکل باید یک Volume دائمی در Railway بسازی و متغیر DATA_DIR را به مسیر Mount آن تنظیم کنی."
        )

    app = (
        Application.builder()
        .token(BOT_TOKEN)
        .base_url(BALE_API_BASE_URL)
        .base_file_url(BALE_API_FILE_URL)
        .post_init(_on_startup)
        .build()
    )
    BOT_INSTANCE = app.bot

    conv_handler = ConversationHandler(
        entry_points=[
            CommandHandler("start", start),
            CallbackQueryHandler(handle_start_upload, pattern="^start_upload$"),
            CallbackQueryHandler(resubmit_barname, pattern="^resubmit_"),
            MessageHandler(filters.Regex("^🚀 بارگزاری مدارک$"), handle_start_upload_button),
        ],
        states={
            WAIT_BARNAME: [
                MessageHandler(filters.Regex("^(🚀 بارگزاری مدارک|📋 راهنما|📦 وضعیت بارنامه|❌ لغو عملیات|🔍 دریافت مستندات|📊 لیست بارنامه‌ها|📈 آمار کلی|🗑 حذف بارنامه|✅ بارنامه‌های تایید شده|🕐 نیازمند بررسی|🏠 منوی اصلی ادمین)$"), reply_keyboard_handler),
                MessageHandler(filters.TEXT & ~filters.COMMAND, receive_barname),
                CallbackQueryHandler(back_to_main, pattern="^back_to_main$"),
            ],
            WAIT_DOCS: [
                MessageHandler(filters.Regex("^❌ لغو عملیات$"), cancel),
                MessageHandler(filters.Regex("^✅ تایید نهایی تمامی مستندات و ارسال به شرکت$"), final_confirm_upload),
                CallbackQueryHandler(final_confirm_upload_inline, pattern="^upload_finalize_inline$"),
                CallbackQueryHandler(cancel_upload_inline, pattern="^upload_cancel_inline$"),
                MessageHandler(filters.Regex("^(🚀 بارگزاری مدارک|📋 راهنما|📦 وضعیت بارنامه|🔍 دریافت مستندات|📊 لیست بارنامه‌ها|📈 آمار کلی|🗑 حذف بارنامه|✅ بارنامه‌های تایید شده|🕐 نیازمند بررسی|🏠 منوی اصلی ادمین)$"), reply_keyboard_handler),
                MessageHandler(filters.PHOTO | filters.Document.ALL, receive_upload_item),
                MessageHandler(filters.TEXT & ~filters.COMMAND, receive_upload_item),
                CallbackQueryHandler(back_to_main, pattern="^back_to_main$"),
            ],
        },
        fallbacks=[
            CommandHandler("cancel", cancel),
            CallbackQueryHandler(back_to_main, pattern="^back_to_main$"),
            MessageHandler(filters.Regex("^❌ لغو عملیات$"), cancel),
        ],
        allow_reentry=True,
    )

    # محافظ ضد آپدیت تکراری — باید زودتر از همه‌ی هندلرهای دیگر اجرا شود (group=-1)
    app.add_handler(TypeHandler(Update, dedupe_update_guard), group=-1)

    app.add_handler(conv_handler)

    # هندلرهای خارج از conversation (ادمین + منو)
    app.add_handler(CallbackQueryHandler(handle_start_upload, pattern="^start_upload$"))
    app.add_handler(CallbackQueryHandler(handle_show_help, pattern="^show_help$"))
    app.add_handler(CallbackQueryHandler(admin_get_docs, pattern="^admin_get$"))
    app.add_handler(CallbackQueryHandler(admin_list_barnames, pattern="^admin_list$"))
    app.add_handler(CallbackQueryHandler(admin_delete_barname, pattern="^admin_delete$"))
    app.add_handler(CallbackQueryHandler(admin_attach_start, pattern="^admin_attach$"))
    app.add_handler(CallbackQueryHandler(admin_stats, pattern="^admin_stats$"))
    app.add_handler(CallbackQueryHandler(admin_approved_list, pattern="^admin_approved_list$"))
    app.add_handler(CallbackQueryHandler(admin_pending_list, pattern="^admin_pending_list$"))
    app.add_handler(CallbackQueryHandler(admin_pending_page, pattern="^admin_pending_page_"))
    app.add_handler(CallbackQueryHandler(admin_open_barname, pattern="^admin_open_"))
    app.add_handler(CallbackQueryHandler(admin_back, pattern="^admin_back$"))
    app.add_handler(CallbackQueryHandler(back_to_main, pattern="^back_to_main$"))
    # هندلرهای بررسی مستندات توسط ادمین (تایید / عدم تایید / تایید با کسری بار)
    app.add_handler(CallbackQueryHandler(review_approve, pattern="^radm_appr_"))
    app.add_handler(CallbackQueryHandler(review_reject_start, pattern="^radm_rej_"))
    app.add_handler(CallbackQueryHandler(review_partial_start, pattern="^radm_part_"))
    app.add_handler(CallbackQueryHandler(review_finalize, pattern="^radm_fin_"))
    app.add_handler(CallbackQueryHandler(driver_partial_accept, pattern="^drv_ok_"))
    app.add_handler(CallbackQueryHandler(driver_partial_reject_start, pattern="^drv_no_"))
    # افزودن عکس/فایل توسط ادمین به یک بارنامه (خارج از فلوی مکالمه‌ی راننده)
    app.add_handler(MessageHandler(filters.PHOTO | filters.Document.ALL, admin_attach_receive_file))
    # هندلر دکمه‌های منوی Reply (باید قبل از admin_text_handler باشه)
    reply_kb_filter = filters.Regex("^(🚀 بارگزاری مدارک|📋 راهنما|📦 وضعیت بارنامه|❌ لغو عملیات|🔍 دریافت مستندات|📊 لیست بارنامه‌ها|📈 آمار کلی|🗑 حذف بارنامه|📎 افزودن عکس/فایل|✅ بارنامه‌های تایید شده|🕐 نیازمند بررسی|🏠 منوی اصلی ادمین)$")
    app.add_handler(MessageHandler(reply_kb_filter, reply_keyboard_handler))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, admin_text_handler))

    logger.info("🤖 ربات پی‌بار v3 شروع به کار کرد...")

    await app.initialize()
    await app.start()
    await app.updater.start_polling(allowed_updates=Update.ALL_TYPES, drop_pending_updates=True)

    logger.info(f"🌐 داشبورد ادمین روی پورت {PORT} در حال اجراست (مسیر: /).")
    uvicorn_config = uvicorn.Config(dashboard_api, host="0.0.0.0", port=PORT, log_level="warning")
    uvicorn_server = uvicorn.Server(uvicorn_config)

    try:
        await uvicorn_server.serve()
    finally:
        await app.updater.stop()
        await app.stop()
        await app.shutdown()


def main():
    asyncio.run(run_app())


if __name__ == "__main__":
    main()
